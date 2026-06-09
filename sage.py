head = False
API_ROOT = "http://139.84.134.18:8002"
DATABASE = "sagepub"
BATCH_SIZE = 10
NUM_PARTS = 6  # split journals into 6 parts for parallel run
SAGE_BASE_URL = "https://journals.sagepub.com"

# When True, images inside article sections are downloaded, JPEG-compressed,
# and embedded as base64. Requires Pillow. Set False to keep URLs only.
SAGE_EMBED_IMAGES = True
SAGE_IMAGE_MAX_DIM = 1600
SAGE_IMAGE_JPEG_QUALITY = 75
SAGE_IMAGE_FETCH_TIMEOUT = 30

# Active Selenium driver exposed to image-embed path (set before calling
# extract_sections_and_access, cleared in finally). Same pattern as rsc.py.
_CURRENT_BROWSER = None

try:
    from PIL import Image as _PIL_Image  # type: ignore
    _HAVE_PIL = True
except Exception:
    _HAVE_PIL = False
import copy
import pycountry
import requests
import json
import os
# import connect_undetected_driver as connect_driver
import time
from difflib import SequenceMatcher
from bs4 import BeautifulSoup
import re
from urllib.parse import urljoin
import urllib.parse
from urllib.parse import parse_qs, unquote
# from status_reporter import report_status
import threading
from selenium.webdriver.common.by import By
# at the top of your file
import ssl
ssl._create_default_https_context = ssl._create_unverified_context
from datetime import datetime
import sys
import multiprocessing
try:
    # Avoid Windows cp1252 UnicodeEncodeError for logs coming from common_code (prints emojis).
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

# Ensure this folder is on sys.path so `utils_vpn.*` imports work
_HERE = os.path.dirname(os.path.abspath(__file__))
if _HERE not in sys.path:
    sys.path.insert(0, _HERE)

from utils_vpn.offline_utils_vpn import (
    create_driver, save_offline, post_journals, save_skipped, save_last_state, save_backup_json,
    read_backup_json, load_last_state, is_driver_alive, safe_get, fetch_and_cache_journals,
    load_journals_from_cache, report_status,
    # load_all_skipped_urls, is_issue_url, parse_issue_url, merge_issue_into_volume_backup,
    force_close_driver,
    extract_country_from_affiliation,
)
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from openpyxl import Workbook, load_workbook


def ensure_sage_journals_split_into_6():
    """
    Fetch all journals from database (API) if needed, then split into 6 part files.
    Uses sage_part_original.txt if present; else fetches from API and saves there, then splits.
    Creates sage_part_1.txt ... sage_part_6.txt (each with "records" = one chunk).
    Idempotent: if sage_part_1.txt exists, skip split.
    """
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    full_path = os.path.join(_script_dir, "sage_part_original.txt")
    def _parts_ok() -> bool:
        """All part files exist and contain at least 1 journal record."""
        for i in range(1, NUM_PARTS + 1):
            p = os.path.join(_script_dir, f"{DATABASE}_part_{i}.txt")
            if not os.path.exists(p):
                return False
            try:
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f) or {}
                recs = data.get("records") or []
                if not recs:
                    return False
            except Exception:
                return False
        return True

    if _parts_ok():
        print(f"✅ Part files already exist and look valid in {_script_dir}. Skip split.")
        return True

    records = []
    if os.path.exists(full_path):
        try:
            with open(full_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            records = data.get("records", [])
            print(f"✅ Loaded {len(records)} journals from {full_path}")
        except Exception as e:
            print(f"⚠️ Error reading {full_path}: {e}")

    if not records:
        print(f"[🌐] Fetching all journals from database ({DATABASE})...")
        page = 1
        while True:
            try:
                resp = requests.get(f"{API_ROOT}/{DATABASE}/journals?page={page}", timeout=60)
                if resp.status_code != 200:
                    break
                data = resp.json()
                chunk = data.get("records", [])
                if not chunk:
                    break
                records.extend(chunk)
                page += 1
            except Exception as e:
                print(f"⚠️ Fetch error: {e}")
                break
        if records:
            try:
                with open(full_path, "w", encoding="utf-8") as f:
                    json.dump({"page": 1, "page_size": len(records), "records": records}, f, indent=2)
                print(f"✅ Saved {len(records)} journals to {full_path}")
            except Exception as e:
                print(f"⚠️ Could not save full file: {e}")

    if not records:
        print("❌ No journals to split.")
        return False

    n = len(records)
    chunk_size = (n + NUM_PARTS - 1) // NUM_PARTS
    for i in range(1, NUM_PARTS + 1):
        start = (i - 1) * chunk_size
        end = min(i * chunk_size, n)
        chunk = records[start:end]
        part_path = os.path.join(_script_dir, f"{DATABASE}_part_{i}.txt")
        try:
            with open(part_path, "w", encoding="utf-8") as f:
                json.dump({"page": i, "page_size": len(chunk), "records": chunk}, f, ensure_ascii=False, indent=2)
            print(f"   Wrote {part_path} ({len(chunk)} journals)")
        except Exception as e:
            print(f"   ⚠️ Failed to write {part_path}: {e}")
    print(f"✅ Split into {NUM_PARTS} parts.")
    return True


def split_journals_into_n_parts(n_parts: int) -> bool:
    """
    Fetch all journals from API (or load sage_part_original.txt), split into n_parts,
    and write sage_part_1.txt … sage_part_N.txt. Always overwrites existing part files.
    Usage: python sage.py split <N>
    """
    _script_dir = os.path.dirname(os.path.abspath(__file__))
    full_path = os.path.join(_script_dir, "sage_part_original.txt")

    records = []
    if os.path.exists(full_path):
        try:
            with open(full_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            records = data.get("records", [])
            print(f"✅ Loaded {len(records)} journals from {full_path}")
        except Exception as e:
            print(f"⚠️ Error reading {full_path}: {e}")

    if not records:
        print(f"[🌐] Fetching all journals from database ({DATABASE})...")
        page = 1
        while True:
            try:
                resp = requests.get(f"{API_ROOT}/{DATABASE}/journals?page={page}", timeout=60)
                if resp.status_code != 200:
                    break
                data = resp.json()
                chunk = data.get("records", [])
                if not chunk:
                    break
                records.extend(chunk)
                page += 1
            except Exception as e:
                print(f"⚠️ Fetch error: {e}")
                break
        if records:
            try:
                with open(full_path, "w", encoding="utf-8") as f:
                    json.dump({"page": 1, "page_size": len(records), "records": records}, f, indent=2)
                print(f"✅ Saved {len(records)} journals to {full_path}")
            except Exception as e:
                print(f"⚠️ Could not save full file: {e}")

    if not records:
        print("❌ No journals to split.")
        return False

    n = len(records)
    print(f"📋 Total journals: {n}  →  splitting into {n_parts} parts")
    chunk_size = (n + n_parts - 1) // n_parts
    for i in range(1, n_parts + 1):
        start = (i - 1) * chunk_size
        end = min(i * chunk_size, n)
        chunk = records[start:end]
        part_path = os.path.join(_script_dir, f"{DATABASE}_part_{i}.txt")
        try:
            with open(part_path, "w", encoding="utf-8") as f:
                json.dump({"page": i, "page_size": len(chunk), "records": chunk}, f, ensure_ascii=False, indent=2)
            print(f"   Wrote {part_path} ({len(chunk)} journals)")
        except Exception as e:
            print(f"   ⚠️ Failed to write {part_path}: {e}")
    print(f"✅ Split complete: {n} journals → {n_parts} parts")
    return True


def ensure_sage_journals_in_article_links_cache() -> str:
    """
    Cache the API `{DATABASE}/journals_in_article_links` response on disk (repo root).
    If file exists, do not call API again.
    """
    cache_path = os.path.join(_HERE, f"{DATABASE}_journals_in_article_links.txt")
    if os.path.exists(cache_path):
        return cache_path
    print(f"[CACHE] Downloading {DATABASE} journals_in_article_links -> {cache_path}")
    resp = requests.get(f"{API_ROOT}/{DATABASE}/journals_in_article_links", timeout=300)
    if resp.status_code != 200:
        raise RuntimeError(f"journals_in_article_links failed: {resp.status_code} {resp.text[:200]}")
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(resp.json(), f, ensure_ascii=False, indent=2)
    return cache_path


def _sage_pairs_index_dir() -> str:
    return os.path.join(_HERE, "journals_in_article_links_index")


def ensure_sage_journals_in_article_links_index(raw_path: str) -> str:
    """
    Build a per-jid index from the cached JSON, without loading it into memory.
    Writes one file per jid: `<sage>/journals_in_article_links_index/<jid>.json`
    Each contains: {"jid": "...", "pairs": [["vol","issue"], ...]}
    """
    index_dir = _sage_pairs_index_dir()
    done_flag = os.path.join(index_dir, ".done")
    if os.path.exists(done_flag):
        return index_dir
    os.makedirs(index_dir, exist_ok=True)

    vol_re = re.compile(r'^\s*"volume"\s*:\s*"([^"]*)"')
    issue_re = re.compile(r'^\s*"issue"\s*:\s*"([^"]*)"')
    jid_re = re.compile(r'^\s*"jid"\s*:\s*"([^"]*)"')

    cur_vol = None
    pairs: list[list[str]] = []
    wrote = 0

    with open(raw_path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            m = vol_re.search(line)
            if m:
                cur_vol = m.group(1)
                continue
            m = issue_re.search(line)
            if m and cur_vol is not None:
                pairs.append([str(cur_vol), str(m.group(1))])
                continue
            m = jid_re.search(line)
            if m:
                jid = m.group(1).strip()
                if jid:
                    out_path = os.path.join(index_dir, f"{jid}.json")
                    try:
                        with open(out_path, "w", encoding="utf-8") as out:
                            json.dump({"jid": jid, "pairs": pairs}, out, ensure_ascii=False)
                        wrote += 1
                    except Exception:
                        pass
                cur_vol = None
                pairs = []
                continue

    with open(done_flag, "w", encoding="utf-8") as f:
        f.write(f"wrote={wrote}\n")
    print(f"[INDEX] Built {DATABASE} journals_in_article_links index: {index_dir} (jids={wrote})")
    return index_dir


def load_extracted_pairs_for_jid(index_dir: str, jid: str) -> set[tuple[str, str]]:
    """Return set of (volume, issue) already present in DB snapshot for this jid."""
    if not jid:
        return set()
    p = os.path.join(index_dir, f"{jid}.json")
    if not os.path.exists(p):
        return set()
    try:
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f) or {}
        pairs = data.get("pairs") or []
        out = set()
        for v, i in pairs:
            v = str(v).strip()
            i = str(i).strip()
            if v and i:
                out.add((v, i))
        return out
    except Exception:
        return set()


def fetch_and_extract_urls(url: str, output_file: str, journal_id: str = None) -> list[str]:
    """
    Fetch or load article URLs. Uniqueness is by (jid, article_url).
    File stores lines "jid\\turl"; we return only URLs for the given journal_id.
    """
    def _read_all_entries():
        if not os.path.exists(output_file) or os.path.getsize(output_file) == 0:
            return []
        entries = []
        with open(output_file, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                if "\t" in line:
                    jid_part, url_part = line.split("\t", 1)
                    entries.append((jid_part.strip(), url_part.strip()))
                else:
                    entries.append((None, line))
        return entries

    if journal_id and os.path.exists(output_file) and os.path.getsize(output_file) > 0:
        entries = _read_all_entries()
        urls_for_jid = [url for jid, url in entries if jid == journal_id]
        if urls_for_jid:
            print(f"[📂] Loaded {len(urls_for_jid)} URLs for jid={journal_id} from {output_file}")
            return urls_for_jid
    elif not journal_id and os.path.exists(output_file) and os.path.getsize(output_file) > 0:
        entries = _read_all_entries()
        urls = [url for _, url in entries]
        if urls:
            print(f"[📂] Loaded {len(urls)} URLs from {output_file}")
            return urls

    try:
        print(f"[🌐] Fetching fresh data from {url} ...")
        response = requests.get(url, timeout=600)
        response.raise_for_status()

        data = response.json()
        article_urls = [item["article_url"] for item in data.get("data", []) if "article_url" in item]

        existing = _read_all_entries()
        other_entries = [(j, u) for j, u in existing if j != journal_id]
        new_entries = [(journal_id or "", url) for url in article_urls] if journal_id else [(None, url) for url in article_urls]
        with open(output_file, "w", encoding="utf-8") as f:
            for jid, url in other_entries:
                f.write(f"{jid}\t{url}\n" if jid else f"{url}\n")
            for jid, url in new_entries:
                f.write(f"{jid}\t{url}\n" if jid else f"{url}\n")

        print(f"[✔] Fetched and saved {len(article_urls)} URLs for jid={journal_id or 'n/a'} to {output_file}")
        return article_urls

    except (requests.RequestException, json.JSONDecodeError, KeyError) as e:
        print(f"[❌] Failed to fetch or parse data from {url}: {e}")
        return []


def _year_from_till_date(till_date):
    """Parse till_date to extract year (e.g. '25 08 2014' or '2014-08-25' -> '2014')."""
    if not till_date:
        return None
    s = str(till_date).strip()
    m = re.match(r"^(\d{4})[-/]", s)
    if m:
        return m.group(1)
    m = re.search(r"(\d{4})\s*$", s)
    if m:
        return m.group(1)
    m = re.search(r"\b(\d{4})\b", s)
    if m:
        return m.group(1)
    return None


def _normalize_volume_issue_map(raw_map):
    """
    Normalize volume_issue_map to one canonical pattern (like new_volume_issue_style.json).
    - Top-level keys: string volume keys (e.g. "38", "2021", "44").
    - Each value: { "volume": str, "year": str, "issues": { issue_num: url } }.
    - Dedupes issues by URL (one issue key per URL; removes duplicate "0" for same volume-only URL).
    Use on load and before save so we always read/store one pattern for future extraction.
    """
    if not raw_map or not isinstance(raw_map, dict):
        return {}
    out = {}
    for vol_key, vol_data in raw_map.items():
        if not isinstance(vol_data, dict):
            continue
        vol_str = str(vol_key)
        volume = str(vol_data.get("volume", vol_str))
        year = str(vol_data.get("year") or "")
        issues_raw = vol_data.get("issues")
        if not isinstance(issues_raw, dict):
            issues_raw = {}
        # Dedupe by URL: keep first issue key per URL (avoids duplicate "0" for same URL)
        seen_urls = set()
        issues = {}
        for issue_num, url in issues_raw.items():
            if not url or url in seen_urls:
                continue
            seen_urls.add(url)
            issues[str(issue_num)] = str(url)
        out[vol_str] = {"volume": volume, "year": year, "issues": issues}
    return out


def _all_issue_urls_from_map(volume_issue_map):
    """Return set of all issue URLs in a volume_issue_map."""
    urls = set()
    if not volume_issue_map or not isinstance(volume_issue_map, dict):
        return urls
    for vol_data in volume_issue_map.values():
        issues = vol_data.get("issues") or {}
        for issue_url in issues.values():
            if issue_url:
                urls.add(issue_url)
    return urls


def _build_to_mine_map(fresh_map, existing_issue_urls, min_year=None, include_year_2025=True):
    """Build map of volumes/issues to mine: new URLs + from min_year onwards (and optionally 2025)."""
    to_mine = {}
    if not fresh_map or not isinstance(fresh_map, dict):
        return to_mine
    for vol_num, vol_data in fresh_map.items():
        year = str(vol_data.get("year") or "")
        if min_year and year.isdigit() and min_year.isdigit():
            if int(year) < int(min_year):
                continue
        issues = vol_data.get("issues") or {}
        new_issues = {}
        for issue_num, issue_url in issues.items():
            if not issue_url:
                continue
            is_new = issue_url not in existing_issue_urls
            is_2025 = include_year_2025 and year == "2025"
            if is_new or is_2025:
                new_issues[issue_num] = issue_url
        if new_issues:
            to_mine[vol_num] = {
                "volume": vol_data.get("volume", vol_num),
                "year": year,
                "issues": new_issues,
            }
    return to_mine

def extract_volumes_issues_sage(SKIPPED_FILE, driver, journal_url, journal_id, min_year=None, existing_map=None):
    """
    Extract SAGE volumes/issues from archive. Only opens decade ranges (dYYYY.yYYYY) where end year >= min_year.
    If existing_map is provided, merges new data into it and saves updated JSON (C:/sage/{jid}/volume_issue_map.json).
    """
    volume_issue_map = copy.deepcopy(existing_map) if existing_map and isinstance(existing_map, dict) else {}

    driver = safe_get(SKIPPED_FILE, driver, journal_url, journal_id)
    print("journal --->", journal_url)
    if not driver:
        return (volume_issue_map, None)

    try:
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        soup = BeautifulSoup(driver.page_source, "html.parser")

        # Pattern from archive: decade tabs have data-url="d1980"; year links have href="/loi/adta/group/d2020.y2025" and data-url="d2020.y2025"
        # Prefer collecting year-level URLs from href (matches actual click: https://journals.sagepub.com/loi/adta/group/d2020.y2025)
        base = "https://journals.sagepub.com"
        year_link_pattern = re.compile(r"^/loi/[^/]+/group/(d\d{4}\.y\d{4})$")
        year_links = soup.find_all("a", href=year_link_pattern)
        if year_links:
            seen_year_urls = set()
            for a in year_links:
                href = a.get("href")
                if not href or href in seen_year_urls:
                    continue
                m = year_link_pattern.match(href)
                if not m:
                    continue
                data_url = m.group(1)
                # data_url is e.g. d2020.y2025 -> year 2025
                ym = re.match(r"d\d{4}\.y(19|20)(\d{2})", data_url)
                year_num = int(ym.group(1) + ym.group(2)) if ym else 0
                if min_year and str(min_year).isdigit() and year_num < int(min_year):
                    continue
                seen_year_urls.add(href)
                year_url = urljoin(base, href)
                print(f"\n🔸 Processing year range: {data_url}  →  {year_url}")
                try:
                    driver = safe_get(SKIPPED_FILE, driver, year_url, journal_id)
                    if not driver:
                        continue
                    decade_soup = BeautifulSoup(driver.page_source, "html.parser")
                except Exception as e:
                    print(f"⚠️ Failed to load year {data_url}: {str(e)}")
                    continue
                # Extract issues from this year page (same as below)
                issue_links = decade_soup.find_all("a", class_="loi__issue__link")
                if not issue_links:
                    issue_links = decade_soup.find_all("a", href=re.compile(r"/toc/[^/]+/\d+"))
                print(f"  Found {len(issue_links)} issue links")
                for issue_a in issue_links:
                    ihref = issue_a.get("href")
                    if not ihref:
                        continue
                    issue_url = urljoin(base, ihref)
                    year_span = issue_a.find("span", class_="loi__issue__cover-date")
                    year_text = year_span.text.strip() if year_span else ""
                    year_match = re.search(r"(19|20)\d{2}", year_text)
                    year = year_match.group(0) if year_match else str(year_num) if year_num else "0"
                    m_vol = re.search(r"/toc/[^/]+/(\d+)(?:/([^/]+))?", ihref)
                    if not m_vol:
                        continue
                    volume_num = m_vol.group(1)
                    issue_num = m_vol.group(2)
                    issue_num = "0" if (issue_num is None or not str(issue_num).strip()) else str(issue_num).strip()
                    if volume_num not in volume_issue_map:
                        volume_issue_map[volume_num] = {"volume": volume_num, "year": year, "issues": {}}
                    if issue_url in set(volume_issue_map[volume_num]["issues"].values()):
                        continue
                    volume_issue_map[volume_num]["issues"][issue_num] = issue_url
                    print(f"Found Vol.{volume_num} Issue {issue_num} ({year}) → {issue_url}")
            if volume_issue_map:
                print(f"\n🎯 Total Volumes Extracted: {len(volume_issue_map)}")
                try:
                    save_backup_json(DATABASE, journal_id, _normalize_volume_issue_map(volume_issue_map), "volume_issue_map.json")
                except Exception:
                    pass
                return (volume_issue_map, driver)

        data_url_links = soup.find_all("a", attrs={"data-url": True})
        print(f"🟢 Found {len(data_url_links)} <a> tags with data-url")

        # If no decade/year links (new layout), try extracting /toc/ links directly from main page
        if not data_url_links:
            direct_links = soup.find_all("a", href=re.compile(r"/toc/[^/]+/\d+"))
            print(f"🟢 No decade links; trying {len(direct_links)} direct /toc/ links on main page")
            for issue_a in direct_links:
                href = issue_a.get("href")
                if not href:
                    continue
                issue_url = urljoin("https://journals.sagepub.com", href)
                m = re.search(r"/toc/[^/]+/(\d+)(?:/([^/]+))?", href)
                if not m:
                    continue
                volume_num = m.group(1)
                issue_num = m.group(2)
                if issue_num is None or issue_num.strip() == "":
                    issue_num = "0"
                else:
                    issue_num = str(issue_num).strip()
                year = volume_num if volume_num.isdigit() and len(volume_num) == 4 else ""
                if volume_num not in volume_issue_map:
                    volume_issue_map[volume_num] = {"volume": volume_num, "year": year, "issues": {}}
                if issue_url in set(volume_issue_map[volume_num]["issues"].values()):
                    continue
                volume_issue_map[volume_num]["issues"][issue_num] = issue_url
                print(f"Found Vol.{volume_num} Issue {issue_num} → {issue_url}")
            if volume_issue_map:
                print(f"\n🎯 Total Volumes Extracted (direct): {len(volume_issue_map)}")
                try:
                    save_backup_json(DATABASE, journal_id, _normalize_volume_issue_map(volume_issue_map), "volume_issue_map.json")
                except Exception:
                    pass
                return (volume_issue_map, driver)

        for a in data_url_links:
            data_url = a.get("data-url")

            if not data_url:
                continue

            match = re.match(r"d(19|20)(\d{2})\.y(19|20)(\d{2})", data_url)
            if not match:
                print(f"⏩ Skipping invalid decade range: {data_url}")
                continue

            # Decade end year (e.g. y2029 -> 2029). Only open if end_year >= min_year
            decade_end_year = int(match.group(3) + match.group(4))
            if min_year and min_year.isdigit():
                if decade_end_year < int(min_year):
                    continue

            decade_url = f"{journal_url}/group/{data_url}"
            print(f"\n🔸 Processing decade range: {data_url}  →  {decade_url}")

            try:
                driver = safe_get(SKIPPED_FILE, driver, decade_url, journal_id)
                # driver.get(decade_url)
                # time.sleep(2)
                # WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                decade_soup = BeautifulSoup(driver.page_source, "html.parser")
            except Exception as e:
                print(f"⚠️ Failed to load decade {data_url}: {str(e)}")
                continue

            # ✅ Extract issues: support old (loi__issue__link) and new layouts; issue can be "9", "9-10", "7-8", or 0 for volume-only
            issue_links = decade_soup.find_all("a", class_="loi__issue__link")
            if not issue_links:
                # Fallback: any link to /toc/ (new style may use different class)
                issue_links = decade_soup.find_all("a", href=re.compile(r"/toc/[^/]+/\d+"))
            print(f"  Found {len(issue_links)} issue links")

            for issue_a in issue_links:
                href = issue_a.get("href")
                if not href:
                    continue

                issue_url = urljoin("https://journals.sagepub.com", href)

                # Extract year from cover-date or from URL (e.g. volume 2021 -> year 2021)
                year_span = issue_a.find("span", class_="loi__issue__cover-date")
                year_text = year_span.text.strip() if year_span else ""
                year_match = re.search(r"(19|20)\d{2}", year_text)
                year = year_match.group(0) if year_match else "0"

                # Volume and issue: support /toc/xxx/38/9, /toc/xxx/38/9-10, /toc/xxx/2021 (volume-only)
                m = re.search(r"/toc/[^/]+/(\d+)(?:/([^/]+))?", href)
                if not m:
                    continue

                volume_num = m.group(1)
                issue_num = m.group(2)
                if issue_num is None or issue_num.strip() == "":
                    issue_num = "0"
                else:
                    issue_num = str(issue_num).strip()

                # Dedupe: don't add same URL twice (avoids repeated links / duplicate "0" entries)
                if volume_num not in volume_issue_map:
                    volume_issue_map[volume_num] = {
                        "volume": volume_num,
                        "year": year if year != "0" else volume_num if volume_num.isdigit() and len(volume_num) == 4 else year,
                        "issues": {}
                    }
                existing_urls = set(volume_issue_map[volume_num]["issues"].values())
                if issue_url in existing_urls:
                    continue
                # Use first available issue key if we already have "0" for same volume (avoid duplicate key)
                issues_dict = volume_issue_map[volume_num]["issues"]
                if issue_num in issues_dict and issues_dict[issue_num] == issue_url:
                    continue
                issues_dict[issue_num] = issue_url

                print(f"Found Vol.{volume_num} Issue {issue_num} ({year}) → {issue_url}")

        print(f"\n🎯 Total Volumes Extracted: {len(volume_issue_map)}")

        if not volume_issue_map:
            print(f"⚠️ No issues found for {journal_id}")

    except Exception as e:
        print(f"❌ Error in extract_volumes_issues_sage for {journal_id}: {str(e)}")
        if not volume_issue_map and existing_map:
            volume_issue_map = copy.deepcopy(existing_map)
    try:
        save_backup_json(DATABASE, journal_id, _normalize_volume_issue_map(volume_issue_map), "volume_issue_map.json")
    except Exception as _:
        pass
    print(volume_issue_map)
    return (volume_issue_map, driver)

def process_articles(
    STATE_FILE,
    SKIPPED_FILE,
    journals,
    start_volume=None,
    start_issue=None,
    start_url=None,
    driver=None,
    resume_journal=None,
    base_url="https://journals.sagepub.com/",
    volume_issue_map=None,
    links_only=False,
    links_queue_file=None,
    links_state_file=None,
):
    """
    Process articles from sage journals.
    links_only=True  → Phase 1: collect article URLs from issue pages only (no article page visits).
                       Links written to links_queue_file (JSONL). Resume state in links_state_file.
    links_only=False → Full mode: visit each article page and extract all data (default).
    """
    # Phase 1 uses its own state file so it doesn't conflict with normal-mode state.
    _state_file = STATE_FILE
    update_volume, update_issue, update_url = None, None, None

    if not driver:
        driver = create_driver(head)

    resumed = True if resume_journal else False
    resume_volume = start_volume
    resume_issue = start_issue
    resume_url = start_url

    for journal in journals:
        journal_id = journal.get("jid")
        journal_url = journal.get("archive_page_url")
        print(f"🔍 Processing {journal_id} → {journal_url}")
        extracted_pairs = journal.get("_extracted_pairs") or set()

        # Unique file per jid: sage_articles_{jid}.txt
        article_file_jid = f"{DATABASE}_articles_{journal_id}.txt"
        article_urls = fetch_and_extract_urls(
            f"http://139.84.134.18:8002/{DATABASE}/all_article_urls?jid={journal_id}",
            article_file_jid,
            journal_id=journal_id,
        )

        if resumed and resume_journal and journal_id != resume_journal:
            print(f"⏭️ Skipping Journal {journal_id} until {resume_journal}")
            continue

        if resumed and resume_journal and journal_id == resume_journal:
            print(f"✅ Resuming from Journal {journal_id}")
            resume_journal = None

        min_year = _year_from_till_date(journal.get("till_date"))

        # ── BOOK: no volumes/issues, paginate through book pages directly ──
        if _is_book_url(journal_url or ""):
            driver = process_book_articles(
                STATE_FILE, SKIPPED_FILE, driver, journal, article_urls,
                base_url=base_url, resumed=resumed, resume_url=resume_url,
                links_only=links_only, links_queue_file=links_queue_file,
                links_state_file=links_state_file,
            )
            # Reset resume flags after book is processed
            resume_volume = resume_issue = resume_url = None
            resumed = False
            continue

        if volume_issue_map is not None and isinstance(volume_issue_map, dict):
            _map = _normalize_volume_issue_map(volume_issue_map)
            print(f"📂 Using provided volume_issue_map ({len(_map)} volumes)")
        else:
            raw_map = read_backup_json(journal_id, DATABASE)
            existing_map = _normalize_volume_issue_map(raw_map) if raw_map else None
            existing_urls = _all_issue_urls_from_map(existing_map) if existing_map else set()
            fresh_map, driver = extract_volumes_issues_sage(
                SKIPPED_FILE, driver, journal_url, journal_id,
                min_year=min_year, existing_map=existing_map,
            )
            to_mine = _build_to_mine_map(fresh_map, existing_urls, min_year=min_year, include_year_2025=True)
            _map = _normalize_volume_issue_map(to_mine if to_mine else fresh_map)
            if to_mine:
                total_new = sum(len(v.get("issues") or {}) for v in to_mine.values())
                print(f"📂 To mine: {len(to_mine)} volumes, {total_new} issues (new / from till_date year)")
        combined_batch = []

        for volume_num, volume_data in _map.items():
            volume = str(volume_data.get("volume", "0"))
            published_year = str(volume_data.get("year", "0"))

            if min_year and published_year.isdigit() and min_year.isdigit():
                if int(published_year) < int(min_year):
                    continue

            # Skip until resume volume
            if resumed and resume_volume and str(volume) != str(resume_volume):
                print(f"⏭️ Skipping Volume {volume} until {resume_volume}")
                continue

            # Once reached resume volume, clear the flag
            if resumed and resume_volume and str(volume) == str(resume_volume):
                print(f"✅ Resuming from Volume {volume}")
                resume_volume = None

            for issue_num, issue_url in volume_data.get("issues", {}).items():
                issue_num = str(issue_num)

                # Skip until resume issue
                if resumed and resume_issue and str(issue_num) != str(resume_issue):
                    print(f"⏭️ Skipping Issue {issue_num} until {resume_issue}")
                    continue

                # Once reached resume issue, clear the flag
                if resumed and resume_issue and str(issue_num) == str(resume_issue):
                    print(f"✅ Resuming from Issue {issue_num}")
                    resume_issue = None

                # Fast skip: if this (volume, issue) is already in DB snapshot, skip loading issue page.
                # But DO NOT skip for current years (2025/2026) because they can update any time.
                try:
                    if extracted_pairs and published_year not in {"2025", "2026"}:
                        if (str(volume).strip(), str(issue_num).strip()) in extracted_pairs:
                            continue
                except Exception:
                    pass

                print(f"📘 Volume: {volume} | Issue: {issue_num} | URL: {issue_url}")

                # After all resume points cleared, turn off resume mode
                if resumed and not (resume_journal or resume_volume or resume_issue or resume_url):
                    print("🟢 Resume complete — processing all subsequent items normally.")
                    resumed = False

                driver = safe_get(SKIPPED_FILE, driver, issue_url, journal_id=journal_id)
                if not driver:
                    continue
                time.sleep(5)
                soup = BeautifulSoup(driver.page_source, "html.parser")

                # -----------------------------
                # Browse Mode (Volume != "0")
                # -----------------------------
                if str(volume) != "0":
                    sections = soup.find_all("div", class_="issue-item__container")
                    for container in sections:
                        title_tags = container.find_all("div", class_="issue-item__title") if container else []
                        for title_tag in title_tags:
                            article_url =""
                            title = ""
                            a = title_tag.find('a')
                            if a:
                                article_url = urljoin(base_url,a["href"])
                                title = title_tag.a.get_text(strip=True)
                            else:
                                continue
                            published_date = ""
                            article_type  = ""
                            header = container.find("div", class_="issue-item__header")
                            if header:
                                spans = header.find_all("span")
                                if len(spans) >= 2:
                                    article_type = spans[1].get_text(strip=True)
                                for sp in spans:
                                    text = sp.get_text(strip=True)
                                    if "First published" in text:
                                        published_date = text.replace("First published", "").strip()
                                        break
                                
                            pdf_url = ""
                            actions = container.find("div", class_="issue-item__actions")
                            if actions:
                                a_pdf = actions.find("a", title="download", href=True)
                                if a_pdf:
                                    pdf_url = urljoin(base_url, a_pdf["href"])
                            
                            
                            # Skip until resume URL
                            if resumed and resume_url and article_url != resume_url:
                                    print(f"⏭️ Skipping article until {resume_url}")
                                    continue

                                # Once reached resume URL, clear it
                            if resumed and resume_url and article_url == resume_url:
                                    print(f"✅ Resuming from article URL {article_url}")
                                    resume_url = None
                                    resumed = False
                            if article_url in article_urls:
                                print("🟡 Already extracted, skipping:", article_url)
                                update_volume, update_issue, update_url = volume, issue_num, article_url
                                save_last_state(STATE_FILE, journal_id, update_url, update_volume, update_issue)
                                continue

                            # ── PHASE 1: links-only mode ──────────────────────────
                            if links_only:
                                link_rec = {
                                    "jid": journal_id,
                                    "article_url": article_url,
                                    "article_title": title,
                                    "article_type": article_type,
                                    "published_date": published_date,
                                    "pdf": pdf_url,
                                    "volume": volume,
                                    "issue": issue_num,
                                    "published_year": published_year,
                                }
                                if links_queue_file:
                                    with open(links_queue_file, "a", encoding="utf-8") as _lf:
                                        _lf.write(json.dumps(link_rec, ensure_ascii=False) + "\n")
                                print(f"  [link] {title[:80]}")
                                update_volume, update_issue, update_url = volume, issue_num, article_url
                                save_last_state(_state_file, journal_id, update_url, update_volume, update_issue)
                                continue
                            # ─────────────────────────────────────────────────────

                            driver = safe_get(SKIPPED_FILE, driver, article_url, journal_id=journal_id)
                            if not driver:
                                    continue
                            time.sleep(4)
                            soup_article = BeautifulSoup(driver.page_source, "html.parser")

                            doi = ""
                            doi_tag = soup_article.find('div', class_="doi")
                            doi = doi_tag.get_text(separator=" ",strip=True) if doi_tag else doi
                            # Abstract
                            abstract = ""
                            abstract_content = soup_article.find('section', id="abstract")
                            if abstract_content:
                                    h2_tag = abstract_content.find('h2')
                                    if h2_tag:
                                            h2_tag.decompose()
                                    abstract = abstract_content.get_text(separator=" ", strip=True) 
                            else:
                                abstract = ""
                            acknowledgment = ""
                            # Acknowledgements
                            acknowledgment_tag = soup_article.find('section', id='acknowledgments')
                            if acknowledgment_tag:
                                h2_tag = acknowledgment_tag.find('h2')
                                if h2_tag:
                                            h2_tag.decompose()
                                acknowledgment = acknowledgment_tag.get_text(strip=True)  
                            else:
                                acknowledgment = ""
                                
                            # Fundings
                            funding = ""
                            funding_tag = soup_article.find('section', id='funding')
                            if funding_tag:
                                # funding_tag.h2.decompose()
                                h2_tag = funding_tag.find('h2')
                                if h2_tag:
                                    h2_tag.decompose()
                                funding = funding_tag.get_text(strip=True)
                            else:
                                funding = ""

                            _page_html = driver.page_source
                            global _CURRENT_BROWSER
                            _CURRENT_BROWSER = driver
                            try:
                                _extras = extract_sections_and_access(_page_html)
                            finally:
                                _CURRENT_BROWSER = None
                            combined_batch.append({
                                        "article_link": {
                                        "jid": journal_id,
                                        "article_title": title,
                                        "article_type": article_type,
                                        "article_url": article_url,
                                        "abstract": abstract,
                                        "pdf": pdf_url,
                                        "doi": doi,
                                        "published_year": published_year,
                                        "published_date": published_date,
                                        "volume": volume,
                                        "issue": issue_num,
                                        "funding": funding,
                                        "funding_information": [],
                                        "acknowledgement": acknowledgment,
                                        "access_type": _extras["access_type"],
                                        "sections": _extras["sections"],
                                        "keywords": _extras["keywords"],
                                        "html": _page_html,
                                        "success": False,
                                        "queue": False
                                    },
                                    "article_data": extract_article_data(_page_html) or []
                            })
                            # Debug: same as OUP / spring_nature
                            article_data_list = combined_batch[-1].get("article_data") or []
                            print("\n" + "=" * 60)
                            print("EXTRACTED VALUES (debug)")
                            print("=" * 60)
                            print(f"  jid                 : {journal_id}")
                            print(f"  article_url         : {article_url}")
                            print(f"  article_title       : {title}")
                            print(f"  article_type        : {article_type}")
                            print(f"  published_year      : {published_year}  |  published_date: {published_date}")
                            print(f"  volume              : {volume}  |  issue: {issue_num}")
                            print(f"  doi                 : {doi}")
                            print(f"  pdf                 : {pdf_url}")
                            print(f"  abstract            : {(abstract or '')[:100]}...")
                            print(f"  funding             : {(funding or '')[:80]}...")
                            print(f"  acknowledgement     : {(acknowledgment or '')[:80]}...")
                            print(f"  article_data (count): {len(article_data_list)} authors")
                            for i, ad in enumerate(article_data_list[:5]):
                                print(f"    author[{i}]       : {ad.get('author_name')} | {ad.get('author_type')} | {ad.get('country')}")
                            if len(article_data_list) > 5:
                                print(f"    ... and {len(article_data_list) - 5} more authors")
                            print("=" * 60 + "\n")
                            print(f"✅ Processed: {title}")

                            # Save progress after each issue
                            update_volume, update_issue, update_url = volume, issue_num, article_url
                            if len(combined_batch) >= BATCH_SIZE:
                                # post_article_links_with_data("sage", combined_batch)
                                save_offline(DATABASE,combined_batch)
                                save_last_state(STATE_FILE, journal_id, update_url, update_volume, update_issue)
                                combined_batch.clear()

        # Post any remaining articles
        if combined_batch:
            # post_article_links_with_data("sage", combined_batch)
            save_offline(DATABASE,combined_batch)
            save_last_state(STATE_FILE, journal_id, update_url, update_volume, update_issue)
            combined_batch.clear()

    return driver


def _is_book_url(url: str) -> bool:
    """True if archive_page_url is a SAGE book (contains /book/ or /doi/book/)."""
    return bool(url and "/book/" in url)


def process_book_articles(
    STATE_FILE, SKIPPED_FILE, driver, journal, article_urls,
    base_url="https://journals.sagepub.com/",
    resumed=False, resume_url=None,
    links_only=False, links_queue_file=None, links_state_file=None,
):
    """
    Process a SAGE book: no volumes/issues. Paginate through all pages of the book,
    extract chapter URLs, visit each to get article_data (same as journal flow).
    links_only=True → Phase 1: save link metadata only, skip article page visits.
    """
    _state_file = STATE_FILE
    journal_id = journal.get("jid")
    book_url = journal.get("archive_page_url")
    print(f"📖 Processing BOOK {journal_id} → {book_url}")

    combined_batch = []
    update_url = None
    page_num = 0

    # Start from page 1 of the book
    current_url = book_url
    if "pageStart" not in current_url:
        sep = "&" if "?" in current_url else "?"
        current_url = f"{current_url}{sep}pageStart=1&pageSize=20"

    # Safety: if resume_url is never found (article removed / pagination changed),
    # we do a second full pass with resume disabled.
    _resume_pass = 1  # pass 1 = try to resume; pass 2 = full scan from start

    while current_url:
        page_num += 1
        print(f"\n📄 Book page {page_num}: {current_url}")

        driver = safe_get(SKIPPED_FILE, driver, current_url, journal_id=journal_id)
        if not driver:
            break
        time.sleep(5)
        soup = BeautifulSoup(driver.page_source, "html.parser")

        # Extract chapter/article URLs from this page
        containers = soup.find_all("div", class_="issue-item__container")
        if not containers:
            print(f"  No articles found on page {page_num}")

        for container in containers:
            title_tags = container.find_all("div", class_="issue-item__title") if container else []
            for title_tag in title_tags:
                a = title_tag.find("a")
                if not a or not a.get("href"):
                    continue
                article_url = urljoin(base_url, a["href"])
                title = a.get_text(strip=True)

                # Published date and article type from header
                published_date = ""
                article_type = ""
                header = container.find("div", class_="issue-item__header")
                if header:
                    spans = header.find_all("span")
                    if len(spans) >= 2:
                        article_type = spans[1].get_text(strip=True)
                    for sp in spans:
                        text = sp.get_text(strip=True)
                        if "First published" in text:
                            published_date = text.replace("First published", "").strip()
                            break

                # PDF URL
                pdf_url = ""
                actions = container.find("div", class_="issue-item__actions")
                if actions:
                    a_pdf = actions.find("a", title="download", href=True)
                    if a_pdf:
                        pdf_url = urljoin(base_url, a_pdf["href"])

                # Skip until resume URL
                if resumed and resume_url and article_url != resume_url:
                    print(f"⏭️ Skipping article until {resume_url}")
                    continue
                if resumed and resume_url and article_url == resume_url:
                    print(f"✅ Resuming from article URL {article_url}")
                    resume_url = None
                    resumed = False

                if article_url in article_urls:
                    print("🟡 Already extracted, skipping:", article_url)
                    update_url = article_url
                    save_last_state(_state_file, journal_id, update_url, "0", "0")
                    continue

                # ── PHASE 1: links-only mode (books) ──────────────────────
                if links_only:
                    link_rec = {
                        "jid": journal_id,
                        "article_url": article_url,
                        "article_title": title,
                        "article_type": article_type or "Book Chapter",
                        "published_date": published_date,
                        "pdf": pdf_url,
                        "volume": "0",
                        "issue": "0",
                        "published_year": "",
                    }
                    if links_queue_file:
                        with open(links_queue_file, "a", encoding="utf-8") as _lf:
                            _lf.write(json.dumps(link_rec, ensure_ascii=False) + "\n")
                    print(f"  [link/book] {title[:80]}")
                    update_url = article_url
                    save_last_state(_state_file, journal_id, update_url, "0", "0")
                    continue
                # ──────────────────────────────────────────────────────────

                # Visit article page for full data
                driver = safe_get(SKIPPED_FILE, driver, article_url, journal_id=journal_id)
                if not driver:
                    continue
                time.sleep(4)
                soup_article = BeautifulSoup(driver.page_source, "html.parser")

                # DOI
                doi = ""
                doi_tag = soup_article.find("div", class_="doi")
                doi = doi_tag.get_text(separator=" ", strip=True) if doi_tag else ""

                # Abstract
                abstract = ""
                abstract_content = soup_article.find("section", id="abstract")
                if abstract_content:
                    h2_tag = abstract_content.find("h2")
                    if h2_tag:
                        h2_tag.decompose()
                    abstract = abstract_content.get_text(separator=" ", strip=True)

                # Acknowledgements
                acknowledgment = ""
                acknowledgment_tag = soup_article.find("section", id="acknowledgments")
                if acknowledgment_tag:
                    h2_tag = acknowledgment_tag.find("h2")
                    if h2_tag:
                        h2_tag.decompose()
                    acknowledgment = acknowledgment_tag.get_text(strip=True)

                # Funding
                funding = ""
                funding_tag = soup_article.find("section", id="funding")
                if funding_tag:
                    h2_tag = funding_tag.find("h2")
                    if h2_tag:
                        h2_tag.decompose()
                    funding = funding_tag.get_text(strip=True)

                _page_html = driver.page_source
                global _CURRENT_BROWSER
                _CURRENT_BROWSER = driver
                try:
                    _extras = extract_sections_and_access(_page_html)
                finally:
                    _CURRENT_BROWSER = None
                combined_batch.append({
                    "article_link": {
                        "jid": journal_id,
                        "article_title": title,
                        "article_type": article_type or "Book Chapter",
                        "article_url": article_url,
                        "abstract": abstract,
                        "pdf": pdf_url,
                        "doi": doi,
                        "published_year": "",
                        "published_date": published_date,
                        "volume": "0",
                        "issue": "0",
                        "funding": funding,
                        "funding_information": [],
                        "acknowledgement": acknowledgment,
                        "access_type": _extras["access_type"],
                        "sections": _extras["sections"],
                        "keywords": _extras["keywords"],
                        "html": _page_html,
                        "success": False,
                        "queue": False,
                    },
                    "article_data": extract_article_data(_page_html) or [],
                })

                article_data_list = combined_batch[-1].get("article_data") or []
                print("\n" + "=" * 60)
                print("EXTRACTED VALUES (book chapter)")
                print("=" * 60)
                print(f"  jid                 : {journal_id}")
                print(f"  article_url         : {article_url}")
                print(f"  article_title       : {title}")
                print(f"  article_type        : {article_type or 'Book Chapter'}")
                print(f"  published_date      : {published_date}")
                print(f"  doi                 : {doi}")
                print(f"  pdf                 : {pdf_url}")
                print(f"  abstract            : {(abstract or '')[:100]}...")
                print(f"  article_data (count): {len(article_data_list)} authors")
                for idx, ad in enumerate(article_data_list[:5]):
                    print(f"    author[{idx}]       : {ad.get('author_name')} | {ad.get('author_type')} | {ad.get('country')}")
                if len(article_data_list) > 5:
                    print(f"    ... and {len(article_data_list) - 5} more authors")
                print("=" * 60 + "\n")
                print(f"✅ Processed: {title}")

                update_url = article_url
                if len(combined_batch) >= BATCH_SIZE:
                    save_offline(DATABASE, combined_batch)
                    save_last_state(STATE_FILE, journal_id, update_url, "0", "0")
                    combined_batch.clear()

        # Check for next page: a.next.hvr-forward.pagination__link
        next_link = soup.select_one("a.next.hvr-forward.pagination__link")
        if next_link and next_link.get("href"):
            current_url = urljoin(base_url, next_link["href"])
        else:
            # Reached the last page of this book
            if resumed and resume_url and _resume_pass == 1:
                # resume_url was never found across the whole book — it was probably
                # removed or the URL changed.  Do a second full pass from page 1.
                print(f"⚠️ Resume URL not found in any page of book {journal_id}: {resume_url}")
                print(f"   Starting full scan from page 1 (pass 2).")
                resumed = False
                resume_url = None
                _resume_pass = 2
                page_num = 0
                current_url = book_url
                if "pageStart" not in current_url:
                    sep = "&" if "?" in current_url else "?"
                    current_url = f"{current_url}{sep}pageStart=1&pageSize=20"
            else:
                print(f"📖 No more pages for book {journal_id}")
                current_url = None

    # Post remaining batch
    if combined_batch:
        save_offline(DATABASE, combined_batch)
        save_last_state(STATE_FILE, journal_id, update_url, "0", "0")
        combined_batch.clear()

    return driver


def extract_article_data(html):
    soup = BeautifulSoup(html, "html.parser")

    def _emails_from_mailto_href(href: str) -> list:
        href = (href or "").strip()
        if not href:
            return []
        if href.lower().startswith("mailto:"):
            href = href[7:]
        href = unquote(href)
        parts = href.split("?", 1)
        primary = (parts[0] or "").strip()
        emails = []
        if primary:
            emails.append(primary)
        if len(parts) == 2 and parts[1]:
            qs = parse_qs(parts[1], keep_blank_values=False)
            for key in ("to", "cc", "bcc"):
                for v in qs.get(key, []):
                    for item in (v or "").replace(";", ",").split(","):
                        item = item.strip()
                        if item:
                            emails.append(item)
        out, seen = [], set()
        for e in emails:
            k = e.lower()
            if k not in seen:
                seen.add(k)
                out.append(e)
        return out

    def _unique_emails(tags):
        """All emails including cc/bcc — used for notes section."""
        out, seen = [], set()
        for a in tags:
            for e in _emails_from_mailto_href(a.get("href", "")):
                k = e.strip().lower()
                if k and k not in seen:
                    seen.add(k)
                    out.append(e.strip())
        return out

    def _primary_emails(tags):
        """Primary (to:) address only — used for author cards to avoid cc contamination."""
        out, seen = [], set()
        for a_tag in tags:
            href = (a_tag.get("href", "") or "").strip()
            if href.lower().startswith("mailto:"):
                href = href[7:]
            em = unquote(href).split("?")[0].strip()
            k = em.lower()
            if em and k not in seen:
                seen.add(k)
                out.append(em)
        return out

    def _email_name_score(email: str, author_name: str) -> int:
        """
        Score 0–100 for how well an email address username matches an author name.
        Used to resolve which author a given email belongs to when there is ambiguity.

        Scoring tiers (first hit wins):
          100 – full name (joined, no spaces) found in username
          100 – full name (space-normalised) equals username
           95 – all name parts (>2 chars) present in username
           90 – first + last name both in username
          ≥90 – SequenceMatcher fuzzy ratio ≥ 90%
           60 – first name (>2 chars) found in username
           50 – last name (>2 chars) found in username
           40 – any name part (>4 chars) found in username
            0 – no match
        """
        username = email.split("@")[0].lower()
        username_norm = re.sub(r'[._\-+]', ' ', username).strip()
        name_lower = author_name.lower().strip()
        parts = name_lower.split()
        first = parts[0] if parts else ""
        last  = parts[-1] if len(parts) > 1 else ""

        if name_lower.replace(' ', '') in username:
            return 100
        if name_lower == username_norm:
            return 100
        long_parts = [p for p in parts if len(p) > 2]
        if long_parts and all(p in username_norm for p in long_parts):
            return 95
        if first and last and first in username_norm and last in username_norm:
            return 90
        ratio = int(SequenceMatcher(None, name_lower, username_norm).ratio() * 100)
        if ratio >= 90:
            return ratio
        if first and len(first) > 2 and first in username_norm:
            return 60
        if last and len(last) > 2 and last in username_norm:
            return 50
        for p in parts:
            if len(p) > 4 and p in username_norm:
                return 40
        return 0

    def _strip_name_prefix_from_address(text: str) -> str:
        """Strip a leading 'Author Name, Institution…' prefix when SAGE embeds the
        author name before their address in a footnote (e.g. pateern6 style)."""
        first_comma = text.find(", ")
        if first_comma <= 0 or first_comma > 60:
            return text
        prefix = text[:first_comma]
        _INST_KW = ('university', 'institute', 'college', 'school', 'hospital',
                    'center', 'department', 'faculty', 'laboratory', 'lab',
                    'gmbh', 'inc', 'ltd', 'corp', 'research', 'agency')
        if any(kw in prefix.lower() for kw in _INST_KW):
            return text  # first segment is an institution name — don't strip
        if len(prefix.split()) > 6:
            return text  # too many words for a personal name
        return text[first_comma + 2:].strip()

    def _name_in_text_score(author_name: str, text: str) -> int:
        """Score 0-100 for how well author_name appears verbatim in a footnote text body."""
        name_lower = author_name.lower().strip()
        text_lower = text.lower()
        parts = name_lower.split()
        long_parts = [p for p in parts if len(p) > 2]
        if not long_parts:
            return 0
        if name_lower in text_lower:
            return 100
        if all(p in text_lower for p in long_parts):
            return 90
        matched = sum(1 for p in long_parts if p in text_lower)
        if matched >= 2:
            return 60
        if matched == 1 and len(long_parts[0]) > 4:
            return 30
        return 0

    def _country_from_affiliation(affiliation_list):
        seen, countries = set(), []
        for aff in affiliation_list:
            c = extract_country_from_affiliation(aff)
            if c and c not in seen:
                seen.add(c)
                countries.append(c)
        return "; ".join(countries)

    # Strip trailing phone/fax block from affiliation text so we can store the
    # address part cleanly without contact details contaminating the field.
    _PHONE_FAX_SUFFIX_RE = re.compile(
        r'(?:[;,\s]+)\b(?:phone|tel(?:ephone)?|fax|mobile|mob|cell)\b.*$',
        re.IGNORECASE,
    )

    def _strip_phone_fax(text: str) -> str:
        return _PHONE_FAX_SUFFIX_RE.sub('', text).strip().rstrip(';,. ')

    # Converts unicode superscript digits/letters to plain ASCII for marker matching
    _SUP_TRANS = str.maketrans("⁰¹²³⁴⁵⁶⁷⁸⁹", "0123456789")

    def _norm_marker(s):
        return s.translate(_SUP_TRANS).strip()

    # Phone number patterns: phone, tel, tel., telephone, mobile, mob., cell + number
    _PHONE_RE = re.compile(
        r'(?:phone|tel\.?|telephone|mobile|mob\.?|cell)[\s.:]*'
        r'([+\d][\d\s\(\)\-\.\/]{4,24})',
        re.IGNORECASE,
    )
    # Tel: X ... Email: Y pairs (for footnotes with multiple authors on one line)
    _TEL_EMAIL_RE = re.compile(
        r'(?:tel\.?|phone|mobile)[\s.:]*([+\d][\d\s\(\)\-\.]{4,20})'
        r'.{0,80}?'
        r'([a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,})',
        re.IGNORECASE,
    )

    def _extract_phones(text):
        phones = []
        for m in _PHONE_RE.finditer(text):
            p = re.sub(r'\s+', ' ', m.group(1)).strip().rstrip('.,;')
            if p and p not in phones:
                phones.append(p)
        return phones

    # ── Parse core-authors-notes: emails + structured footnote entries ─────────
    # Handles three SAGE marker styles:
    #   1. <sup>1</sup>University text          → marker "1"
    #   2. <div class="label">1</div>...        → marker "1"  (pateern3)
    #   3. id="corresp1-..."  (no sup/label)    → marker "corresp1" (pateern4)
    corresponding_emails: set = set()
    footnote_entries = []   # list of {marker, text, emails, phones, tel_email_pairs}

    notes_section = soup.select_one("section.core-authors-notes")
    if notes_section:
        for em in _unique_emails(notes_section.select("a[href^='mailto:']")):
            corresponding_emails.add(em.lower())

        for fn_elem in notes_section.find_all(attrs={"role": "doc-footnote"}):
            # ── Determine marker ──────────────────────────────────────────────
            sup_tag   = fn_elem.find("sup")
            label_tag = fn_elem.find("div", class_="label")
            if sup_tag:
                raw_marker = _norm_marker(sup_tag.get_text(strip=True))
            elif label_tag:
                raw_marker = _norm_marker(label_tag.get_text(strip=True))
            else:
                fn_id  = fn_elem.get("id", "")
                m_id   = re.match(r'^(corresp\d*)', fn_id)
                raw_marker = m_id.group(1) if m_id else ""

            # ── Clean text: strip marker elements + mailto links ─────────────
            clone = BeautifulSoup(str(fn_elem), "html.parser")
            for t in clone.find_all("sup"):
                t.decompose()
            for t in clone.find_all("div", class_="label"):
                t.decompose()
            for t in clone.find_all("a", href=re.compile(r'^mailto:', re.I)):
                t.decompose()
            fn_text = clone.get_text(" ", strip=True).lstrip(".,;: ")

            # Skip pure ORCID / header lines
            if "orcid" in fn_text.lower() or fn_text.lower().startswith("http"):
                continue

            fn_emails = _unique_emails(fn_elem.select("a[href^='mailto:']"))
            fn_phones = _extract_phones(fn_text)

            # Extract (tel, email) pairs from text BEFORE removing mailto links so the
            # email addresses are still present for matching.
            clone_raw = BeautifulSoup(str(fn_elem), "html.parser")
            for t in clone_raw.find_all("sup"):
                t.decompose()
            for t in clone_raw.find_all("div", class_="label"):
                t.decompose()
            fn_text_raw = clone_raw.get_text(" ", strip=True)
            fn_tel_email = [
                (re.sub(r'\s+', ' ', m.group(1)).strip().rstrip('.,;'), m.group(2).strip())
                for m in _TEL_EMAIL_RE.finditer(fn_text_raw)
            ]

            # One raw_marker may be "1,2" → expand
            for part in (re.split(r"[,\s]+", raw_marker) if raw_marker else [""]):
                part = part.strip()
                footnote_entries.append({
                    "marker":         part,
                    "text":           fn_text,
                    "emails":         fn_emails,
                    "phones":         fn_phones,
                    "tel_email_pairs": fn_tel_email,
                })

    # ── Author blocks (two SAGE layouts) ──────────────────────────────────────
    author_sections = soup.select("section.core-authors > div[data-expandable='item']")
    if not author_sections:
        author_sections = soup.select("section.core-authors > div[property='author'][typeof='Person']")

    raw = []
    for div in author_sections:
        # Name — strip * marker (used by SAGE to flag corresponding author)
        # Also strip <sup> (footnote markers) from h4 before extracting text
        name_tag = div.select_one("h4")
        if name_tag:
            h4_clone = BeautifulSoup(str(name_tag), "html.parser")
            for t in h4_clone.find_all("sup"):
                t.decompose()
            name = h4_clone.get_text(" ", strip=True)
        else:
            given  = div.select_one("span[property='givenName']")
            family = div.select_one("span[property='familyName']")
            name   = " ".join(p.get_text(" ", strip=True) for p in [given, family] if p).strip()
        has_star = "*" in name
        name = name.replace("*", "").strip()

        # ORCID
        orcid_tag = div.select_one("a.orcid-id[href*='orcid.org']")
        orcid_id  = orcid_tag["href"].strip() if orcid_tag else ""

        # Author card emails — primary address only (cc/bcc excluded to prevent cross-author pollution)
        author_card_emails = _primary_emails(div.select("a[href^='mailto:']"))

        # Superscript markers on this author card (e.g. <sup>1,2</sup> or <sup class="xref">)
        author_markers = []
        for sup in div.select("sup"):
            for part in re.split(r"[,\s]+", _norm_marker(sup.get_text(strip=True))):
                part = part.strip()
                if part:
                    author_markers.append(part)

        # Affiliations from div.affiliations (structured)
        aff_tags = div.select("div.affiliations span[property='name']")
        affiliation_list = [t.get_text(" ", strip=True) for t in aff_tags] if aff_tags else []
        if not affiliation_list:
            aff_div = div.select_one("div.affiliations")
            if aff_div:
                txt = aff_div.get_text(" ", strip=True)
                if txt:
                    affiliation_list = [txt]

        # Match footnote entries for this author:
        #   marker_match → add affiliation text + phones
        #   email_match  → phones only (avoid duplicating address already in div.affiliations)
        seen_affs  = set(affiliation_list)
        author_phones = []
        author_email_set = {e.lower() for e in author_card_emails}

        fn_matched_ids: set = set()

        for fn in footnote_entries:
            marker_match = bool(fn["marker"]) and fn["marker"] in author_markers
            email_match  = any(fe.lower() in author_email_set for fe in fn["emails"])

            if marker_match:
                fn_matched_ids.add(id(fn))
                # Affiliation: strip trailing phone/fax, then skip contact-only blocks
                fn_aff = _strip_phone_fax(fn["text"])
                _is_contact_block = (
                    not fn_aff
                    or len(fn_aff) <= 8
                    or fn_aff.lower().startswith("corresponding")
                )
                if not _is_contact_block and fn_aff not in seen_affs:
                    affiliation_list.append(fn_aff)
                    seen_affs.add(fn_aff)
                # Phone: use per-author (tel, email) pairs when available.
                # With multiple pairs, prefer name-score matching over card-email
                # matching — card email may point to another author (e.g. ?cc= pattern).
                if fn["tel_email_pairs"]:
                    if len(fn["tel_email_pairs"]) == 1:
                        author_phones.append(fn["tel_email_pairs"][0][0])
                    else:
                        best_ph, best_sc = "", 0
                        for ph, em in fn["tel_email_pairs"]:
                            sc = _email_name_score(em, name)
                            if sc > best_sc:
                                best_sc, best_ph = sc, ph
                        if best_ph and best_sc > 0:
                            author_phones.append(best_ph)
                        else:
                            by_email = [ph for ph, em in fn["tel_email_pairs"]
                                        if em.lower() in author_email_set]
                            author_phones.extend(by_email if by_email else fn["phones"])
                else:
                    author_phones.extend(fn["phones"])

            if email_match and not marker_match:
                fn_matched_ids.add(id(fn))
                # Affiliation: strip name prefix + trailing phone/fax, replace shorter entry
                fn_aff = _strip_phone_fax(_strip_name_prefix_from_address(fn["text"]))
                _is_contact_block_em = (
                    not fn_aff
                    or len(fn_aff) <= 8
                    or fn_aff.lower().startswith("corresponding")
                )
                if not _is_contact_block_em and fn_aff not in seen_affs:
                    # Replace shorter existing entry in-place if fn_aff subsumes its keywords
                    replaced_idx = None
                    for i, existing in enumerate(affiliation_list):
                        exist_keys = [w for w in existing.lower().split()
                                      if len(w) > 4 and w.isalpha()]
                        if exist_keys and all(w in fn_aff.lower() for w in exist_keys):
                            seen_affs.discard(existing)
                            affiliation_list[i] = fn_aff
                            seen_affs.add(fn_aff)
                            replaced_idx = i
                            break
                    if replaced_idx is None:
                        affiliation_list.append(fn_aff)
                        seen_affs.add(fn_aff)
                # Phone via (tel, email) pairs first
                if fn["tel_email_pairs"]:
                    if len(fn["tel_email_pairs"]) == 1:
                        author_phones.append(fn["tel_email_pairs"][0][0])
                    else:
                        best_ph, best_sc = "", 0
                        for ph, em in fn["tel_email_pairs"]:
                            sc = _email_name_score(em, name)
                            if sc > best_sc:
                                best_sc, best_ph = sc, ph
                        if best_ph and best_sc > 0:
                            author_phones.append(best_ph)
                        else:
                            by_email = [ph for ph, em in fn["tel_email_pairs"]
                                        if em.lower() in author_email_set]
                            author_phones.extend(by_email if by_email else fn["phones"])
                else:
                    author_phones.extend(fn["phones"])

        # ── Name-based matching: footnotes not reached by marker/email ────────────
        # Handles cases where author name appears in footnote body (e.g. bio footnotes,
        # or address footnotes that start with the author's name).
        for fn in footnote_entries:
            if id(fn) in fn_matched_ids:
                continue
            if _name_in_text_score(name, fn["text"]) < 70:
                continue
            fn_aff = _strip_phone_fax(fn["text"])
            # Bio text: starts with author's name or contains "is a/an/was/worked as"
            _is_bio = (
                fn_aff.lower().startswith(name.lower()[:min(8, len(name))])
                or bool(re.search(
                    r'\b(?:is an?\s|are an?\s|was an?\s|worked? as\b|serves? as\b)',
                    fn_aff, re.I,
                ))
            )
            _is_contact = (
                not fn_aff
                or len(fn_aff) <= 8
                or fn_aff.lower().startswith("corresponding")
            )
            if not _is_bio and not _is_contact and fn_aff not in seen_affs:
                affiliation_list.append(fn_aff)
                seen_affs.add(fn_aff)
            if fn["phones"]:
                author_phones.extend(fn["phones"])

        # Also extract phones embedded directly in affiliation text
        for aff in affiliation_list:
            author_phones.extend(_extract_phones(aff))

        # Deduplicate phones
        seen_ph, deduped_phones = set(), []
        for p in author_phones:
            pk = re.sub(r'[\s\-\.\(\)]', '', p)
            if pk not in seen_ph:
                seen_ph.add(pk)
                deduped_phones.append(p)

        affiliation = "; ".join(a.rstrip(' ,.;') for a in affiliation_list if a.strip())
        country     = _country_from_affiliation(affiliation_list)
        phone       = "; ".join(deduped_phones)

        raw.append({
            "author_name": name,
            "has_star":    has_star,
            "orcid_id":    orcid_id,
            "emails":      author_card_emails,
            "_markers":    author_markers,
            "affiliation": affiliation,
            "country":     country,
            "phone":       phone,
        })

    if not raw:
        return []

    # ── Deduplicate emails across authors via name-score matching ─────────────
    # Rules:
    #   sole claimant (any score)   → stays with that author
    #   multiple claimants, clear winner (score > all others) → winner keeps it
    #   multiple claimants, score 0 for all OR tie            → orphan record (author_name="")
    email_claims: dict = {}   # email.lower() → [(score, author_idx, original_case)]
    for idx, a in enumerate(raw):
        for em in a["emails"]:
            key = em.lower()
            score = _email_name_score(em, a["author_name"])
            email_claims.setdefault(key, []).append((score, idx, em))

    email_winner: dict = {}   # email.lower() → winning author_idx, or None = orphan
    orphan_emails: list = []  # original-case emails that become standalone records

    for key, claims in email_claims.items():
        max_score = max(s for s, _, _ in claims)
        top = [c for c in claims if c[0] == max_score]

        if len(claims) == 1:
            # Sole claimant — keep regardless of score (opaque username is fine)
            email_winner[key] = claims[0][1]
        elif max_score == 0 or len(top) > 1:
            # No clear owner → orphan
            email_winner[key] = None
            orphan_emails.append(top[0][2])   # preserve original case
        else:
            email_winner[key] = top[0][1]     # unique highest scorer wins

    for idx, a in enumerate(raw):
        a["emails"] = [em for em in a["emails"]
                       if email_winner[em.lower()] == idx]

    # ── Email recovery: authors who lost all emails try to get one from footnotes ─
    # Handles cards where the primary mailto pointed to another author (e.g. ?cc= pattern).
    # Only picks emails not already held by another author; requires positive name-score.
    all_claimed = {e.lower() for b in raw for e in b["emails"]}
    for a in raw:
        if a["emails"]:
            continue
        for fn in footnote_entries:
            marker_match = bool(fn["marker"]) and fn["marker"] in a.get("_markers", [])
            name_sc = _name_in_text_score(a["author_name"], fn["text"])
            if not (marker_match or name_sc >= 70) or not fn["emails"]:
                continue
            unclaimed = [e for e in fn["emails"] if e.lower() not in all_claimed]
            candidates = unclaimed if unclaimed else fn["emails"]
            best = max(candidates, key=lambda e: _email_name_score(e, a["author_name"]))
            if _email_name_score(best, a["author_name"]) > 0:
                a["emails"] = [best]
                all_claimed.add(best.lower())
                break

    # ── Determine author_type ─────────────────────────────────────────────────
    # Rule 1: single author → always Corresponding
    # Rule 2: any name has * → star = Corresponding, rest = Co-Author
    # Rule 3: email matches core-authors-notes → Corresponding, rest = Co-Author
    # Rule 3b: exactly one author has email (after dedup+recovery) → Corresponding
    # Rule 4: none of the above → all = "Author"
    star_set = {a["author_name"] for a in raw if a["has_star"]}

    if len(raw) == 1:
        for a in raw:
            a["author_type"] = "Corresponding"
    elif star_set:
        for a in raw:
            a["author_type"] = "Corresponding" if a["author_name"] in star_set else "Co-Author"
    elif corresponding_emails:
        for a in raw:
            matched = any(e.lower() in corresponding_emails for e in a["emails"])
            a["author_type"] = "Corresponding" if matched else "Co-Author"
    elif sum(1 for a in raw if a["emails"]) == 1:
        for a in raw:
            a["author_type"] = "Corresponding" if a["emails"] else "Co-Author"
    else:
        for a in raw:
            a["author_type"] = "Author"

    # ── Build output — one record per email, Pattern 2 format ─────────────────
    authors_list = []
    for a in raw:
        emails = a["emails"]
        if len(emails) > 1:
            # Multiple emails → one record per email
            for em in emails:
                authors_list.append({
                    "author_name": a["author_name"],
                    "orcid_id":    a["orcid_id"],
                    "email":       em,
                    "affiliation": a["affiliation"],
                    "country":     a["country"],
                    "phone":       a["phone"],
                    "author_type": a["author_type"],
                })
        else:
            authors_list.append({
                "author_name": a["author_name"],
                "orcid_id":    a["orcid_id"],
                "email":       emails[0] if emails else "",
                "affiliation": a["affiliation"],
                "country":     a["country"],
                "phone":       a["phone"],
                "author_type": a["author_type"],
            })

    # ── Orphan records (no clear owner after name-score dedup) ───────────────
    for em in orphan_emails:
        authors_list.append({
            "author_name": "",
            "orcid_id":    "",
            "email":       em,
            "affiliation": "",
            "country":     "",
            "phone":       "",
            "author_type": "",
        })

    return authors_list


# ─────────────────────────────────────────────────────────────────────────────
# Section extraction — canonical name mapping + structured blocks
# (mirrors rsc.py approach, adapted for SAGE HTML conventions)
# ─────────────────────────────────────────────────────────────────────────────

_SECTION_CANONICAL_MAP = [
    (("graphical abstract", "visual abstract", "pictorial abstract"), "GRAPHICAL_ABSTRACT"),
    (("abstract",), "ABSTRACT"),
    (("keywords", "key words", "index terms", "mesh terms", "author keywords"), "KEYWORDS"),
    (("introduction", "background and introduction", "overview"), "INTRODUCTION"),
    (("strengths and limitations", "limitations of the study", "study limitations",
      "limitations"), "LIMITATIONS"),
    (("challenges and future directions", "challenges and perspectives",
      "future work", "future directions", "future research", "future perspectives",
      "perspectives", "challenges and opportunities", "open problems",
      "remaining challenges"), "FUTURE_DIRECTIONS"),
    (("concluding remarks", "final remarks", "closing remarks",
      "conclusions and outlook", "conclusions and perspectives",
      "conclusions", "conclusion", "summary", "outlook"), "CONCLUSION"),
    (("background", "literature review", "related work", "prior work",
      "state of the art", "theoretical background"), "BACKGROUND"),
    (("materials and methods", "materials & methods", "patients and methods",
      "subjects and methods", "methodology", "methods", "experimental section",
      "experimental procedures", "experimental", "study design", "survey methodology",
      "research design", "approach", "procedure"), "MATERIALS_METHODS"),
    (("experimental results", "results and data", "results", "findings",
      "observations", "outcomes"), "RESULTS"),
    (("results and discussion",), "RESULTS"),
    (("discussion", "interpretation", "implications", "commentary",
      "evaluation", "analysis"), "DISCUSSION"),
    (("case report", "case description", "case presentation", "patient information",
      "case summary", "case history", "presenting concerns"), "CASE_PRESENTATION"),
    (("acknowledgements", "acknowledgments", "acknowledgement", "acknowledgment"),
     "ACKNOWLEDGMENTS"),
    (("funding information", "funding sources", "financial support",
      "grant information", "funding", "support"), "FUNDING"),
    (("conflict of interest", "conflicts of interest", "competing interests",
      "declaration of interest", "coi statement", "disclosures",
      "competing declarations"), "CONFLICT_OF_INTEREST"),
    (("ethics statement", "irb approval", "ethical considerations",
      "ethical approval", "ethics approval", "institutional review"), "ETHICS_STATEMENT"),
    (("data availability statement", "data availability", "data statement",
      "data sharing", "open data"), "DATA_AVAILABILITY"),
    (("electronic supplementary material", "supplementary materials",
      "supplementary material", "supplementary data", "supporting information",
      "additional information", "online resources", "appendix"), "SUPPLEMENTARY"),
    (("author contributions", "authors' contributions", "credit author statement",
      "credit authorship contribution statement", "author roles",
      "contributors", "contributions"), "AUTHOR_CONTRIBUTION"),
    (("list of abbreviations", "abbreviations", "nomenclature",
      "glossary", "symbols", "notation"), "ABBREVIATIONS"),
    (("references", "bibliography", "works cited", "literature cited",
      "citations"), "REFERENCES"),
]

# Known section id → canonical name (for named SAGE sections outside bodymatter)
_SAGE_SECTION_ID_MAP = {
    "abstract": "ABSTRACT",
    "funding": "FUNDING",
    "acknowledgments": "ACKNOWLEDGMENTS",
    "acknowledgements": "ACKNOWLEDGMENTS",
    "conflict": "CONFLICT_OF_INTEREST",
    "data-availability": "DATA_AVAILABILITY",
    "bibliography": "REFERENCES",
    "orcid": None,  # skip
}


def _sage_normalize_heading(raw: str) -> str:
    """Map a raw heading like '2.1 Materials and Methods' → 'MATERIALS_METHODS'."""
    if not raw:
        return "OTHERS"
    t = re.sub(r"^\s*\d+(?:\.\d+)*\.?\s*", "", raw).strip().lower()
    t = re.sub(r"[.:;]+$", "", t).strip()
    if not t:
        return "OTHERS"
    for variants, canonical in _SECTION_CANONICAL_MAP:
        if t in variants:
            return canonical
    return "OTHERS"


def _sage_clean(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "")).strip()


def _sage_fetch_image(url: str) -> dict:
    """Download + JPEG-compress an image. Browser path preferred, then requests+Pillow."""
    # Path 1: browser fetch (same session, no CDN blocks)
    if _CURRENT_BROWSER is not None:
        try:
            _JS = """
var cb = arguments[arguments.length-1];
fetch(arguments[0],{credentials:'include',cache:'force-cache'})
  .then(function(r){if(!r.ok){cb({err:'HTTP '+r.status});return;}
    var mime=r.headers.get('content-type')||'image/jpeg';
    return r.arrayBuffer().then(function(buf){
      var b=new Uint8Array(buf),s='';
      for(var i=0;i<b.length;i+=0x8000)s+=String.fromCharCode.apply(null,b.subarray(i,i+0x8000));
      cb({data:btoa(s),mime:mime.split(';')[0].trim(),bytes:b.length});})})
  .catch(function(e){cb({err:String(e)});});"""
            _CURRENT_BROWSER.set_script_timeout(SAGE_IMAGE_FETCH_TIMEOUT)
            res = _CURRENT_BROWSER.execute_async_script(_JS, url)
            if isinstance(res, dict) and res.get("data") and not res.get("err"):
                if _HAVE_PIL:
                    return _sage_compress_jpeg(res["data"]) or res
                return res
        except Exception:
            pass
    # Path 2: direct HTTPS + Pillow
    if not _HAVE_PIL:
        return {}
    try:
        import base64
        from io import BytesIO
        sess = requests.Session()
        sess.headers["User-Agent"] = (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/120.0 Safari/537.36"
        )
        r = sess.get(url, timeout=SAGE_IMAGE_FETCH_TIMEOUT, stream=False)
        if r.status_code != 200:
            return {}
        from io import BytesIO
        img = _PIL_Image.open(BytesIO(r.content))
        return _sage_process_pil(img)
    except Exception:
        return {}


def _sage_compress_jpeg(b64: str) -> dict:
    """Re-encode base64 image as downscaled JPEG."""
    if not _HAVE_PIL:
        return {}
    try:
        import base64
        from io import BytesIO
        raw = base64.b64decode(b64)
        img = _PIL_Image.open(BytesIO(raw))
        return _sage_process_pil(img)
    except Exception:
        return {}


def _sage_process_pil(img) -> dict:
    """Downscale and JPEG-encode a PIL Image, return base64 dict."""
    import base64
    from io import BytesIO
    if img.mode in ("RGBA", "LA", "P"):
        bg = _PIL_Image.new("RGB", img.size, (255, 255, 255))
        bg.paste(img.convert("RGBA"), mask=img.convert("RGBA").split()[-1])
        img = bg
    elif img.mode != "RGB":
        img = img.convert("RGB")
    if max(img.size) > SAGE_IMAGE_MAX_DIM:
        scale = SAGE_IMAGE_MAX_DIM / float(max(img.size))
        img = img.resize(
            (int(img.size[0] * scale), int(img.size[1] * scale)),
            _PIL_Image.LANCZOS,
        )
    buf = BytesIO()
    img.save(buf, format="JPEG", quality=SAGE_IMAGE_JPEG_QUALITY, optimize=True, progressive=True)
    data = buf.getvalue()
    return {
        "data": base64.b64encode(data).decode("ascii"),
        "mime": "image/jpeg",
        "bytes": len(data),
        "width": img.size[0],
        "height": img.size[1],
    }


def _sage_img_block(img_tag) -> dict:
    """Build an img block from a BS4 <img> tag; embed bytes when possible."""
    src = (img_tag.get("src") or "").strip()
    if not src or "LoadingBackGround" in src:
        src = (img_tag.get("data-original") or img_tag.get("data-src") or src).strip()
    if not src:
        return {}
    abs_src = urljoin(SAGE_BASE_URL, src)
    block = {"type": "img", "src": abs_src, "alt": (img_tag.get("alt") or "").strip()}
    if SAGE_EMBED_IMAGES and (_CURRENT_BROWSER is not None or _HAVE_PIL):
        block.update(_sage_fetch_image(abs_src))
    return block


def _sage_table_to_block(tbl) -> dict:
    """Extract a table element into a tbl block with rows + raw HTML."""
    rows = []
    for tr in tbl.find_all("tr"):
        cells = [c.get_text(" ", strip=True) for c in tr.find_all(["th", "td"])]
        if cells:
            rows.append(cells)
    block = {"type": "tbl", "rows": rows, "html": str(tbl)}
    cap = tbl.find("caption")
    if cap:
        cap_text = _sage_clean(cap.get_text(" ", strip=True))
        if cap_text:
            block["caption"] = cap_text
    return block


def _sage_figure_to_blocks(fig) -> list:
    """Convert a SAGE <figure> element to one or more blocks."""
    # Extract caption (remove the <span class="heading"> label like "Figure 1")
    caption = ""
    cap_tag = fig.find("figcaption")
    if cap_tag:
        heading_span = cap_tag.find("span", class_="heading")
        if heading_span:
            heading_span.decompose()
        caption = _sage_clean(cap_tag.get_text(" ", strip=True))

    # Table figure
    tbl = fig.find("table")
    if tbl:
        block = _sage_table_to_block(tbl)
        if caption:
            block["caption"] = caption
        return [block]

    # Image figure
    img_tag = fig.find("img")
    if img_tag:
        img_block = _sage_img_block(img_tag)
        if img_block:
            if caption:
                return [{"type": "fig", "image": img_block, "caption": caption}]
            return [img_block]

    if caption:
        return [{"type": "p", "text": caption}]
    return []


def _sage_element_to_blocks(elem) -> list:
    """Convert one BS4 element to typed blocks (SAGE HTML conventions)."""
    if not hasattr(elem, "name") or not elem.name:
        return []
    name = elem.name.lower()

    # Skip chrome elements
    if name in ("button", "script", "style", "nav"):
        return []

    # Sub-headings (h3, h4, h5)
    if name in ("h3", "h4", "h5"):
        text = _sage_clean(elem.get_text(" ", strip=True))
        return [{"type": "h", "level": int(name[1]), "text": text}] if text else []

    # Skip section-level headings (h2) inside nested sections
    if name == "h2":
        text = _sage_clean(elem.get_text(" ", strip=True))
        return [{"type": "h", "level": 2, "text": text}] if text else []

    # Paragraph: <div role="paragraph">
    if name == "div" and elem.get("role") == "paragraph":
        text = _sage_clean(elem.get_text(" ", strip=True))
        return [{"type": "p", "text": text}] if text else []

    # List: <div role="list">
    if name == "div" and elem.get("role") == "list":
        items = []
        for item in elem.select("div[role='listitem']"):
            content = item.find("div", class_="content") or item
            text = _sage_clean(content.get_text(" ", strip=True))
            if text:
                items.append(text)
        return [{"type": "list", "ordered": False, "items": items}] if items else []

    # Figure element
    if name == "figure":
        return _sage_figure_to_blocks(elem)

    # Table element (outside figure)
    if name == "table":
        return [_sage_table_to_block(elem)]

    # Figure-wrap div (contains button + figure)
    if name == "div":
        cls = " ".join(elem.get("class") or [])
        if "figure-wrap" in cls:
            fig = elem.find("figure")
            return _sage_figure_to_blocks(fig) if fig else []
        # Transparent div — recurse into children
        out = []
        for child in elem.children:
            out.extend(_sage_element_to_blocks(child))
        return out

    # Nested section — recurse
    if name == "section":
        out = []
        for child in elem.children:
            out.extend(_sage_element_to_blocks(child))
        return out

    # Any other inline/block element — recurse
    out = []
    for child in elem.children:
        out.extend(_sage_element_to_blocks(child))
    return out


def _sage_collect_blocks(sec_elem) -> list:
    """Walk a SAGE section element and collect blocks, skipping its own heading."""
    heading_consumed = False
    out = []
    for child in sec_elem.children:
        if not hasattr(child, "name") or not child.name:
            continue
        name = child.name.lower()
        if not heading_consumed and name in ("h2", "h3", "h4"):
            heading_consumed = True
            continue
        out.extend(_sage_element_to_blocks(child))
    return out


def _sage_blocks_to_text(blocks: list) -> str:
    """Flatten blocks to a single plain-text string."""
    parts = []
    for b in blocks or []:
        t = b.get("type")
        if t in ("p", "h", "quote"):
            txt = b.get("text") or ""
            if txt:
                parts.append(txt)
        elif t == "list":
            for item in b.get("items") or []:
                if item:
                    parts.append(f"- {item}")
        elif t == "tbl":
            for row in b.get("rows") or []:
                parts.append(" | ".join(str(c) for c in row if c))
        elif t == "img":
            cap = b.get("alt") or ""
            if cap:
                parts.append(f"[Image: {cap}]")
        elif t == "fig":
            cap = b.get("caption") or (b.get("image") or {}).get("alt") or ""
            if cap:
                parts.append(f"[Figure: {cap}]")
    return _sage_clean("\n\n".join(parts))


def _sage_build_content(blocks: list) -> dict:
    """Wrap blocks into section payload. Text-only when all blocks are paragraphs."""
    text = _sage_blocks_to_text(blocks)
    has_structure = any(b.get("type") != "p" for b in (blocks or []))
    if has_structure:
        return {"blocks": blocks, "text": text}
    return {"text": text}


def extract_sections_and_access(html: str) -> dict:
    """
    Extract structured article sections, access type, and keywords from SAGE HTML.

    sections dict uses canonical keys (ABSTRACT, INTRODUCTION, MATERIALS_METHODS, …)
    matching the RSC / science_direct format. Each section value is:
      • a plain string for TITLE
      • {"text": "…"} for text-only sections
      • {"blocks": […], "text": "…"} when the section contains images / tables / lists

    Images are downloaded and base64-embedded when _CURRENT_BROWSER is set
    (set it before calling this function, clear in finally — same as rsc.py).
    """
    soup = BeautifulSoup(html, "html.parser")

    # ── Access type ────────────────────────────────────────────────────────
    access_type = "unknown"
    if (soup.find(class_=re.compile(r"icon-open_access", re.I))
            or soup.find("a", href=re.compile(r"creativecommons\.org", re.I))):
        access_type = "open_access"
    elif soup.find(class_=re.compile(r"denial-block", re.I)):
        access_type = "restricted"

    # ── Keywords ───────────────────────────────────────────────────────────
    kw_meta = soup.find("meta", attrs={"name": "keywords"})
    keywords = []
    if kw_meta and kw_meta.get("content"):
        keywords = [k.strip() for k in kw_meta["content"].split(",") if k.strip()]

    # ── Sections ───────────────────────────────────────────────────────────
    out = {}
    others = []

    def _merge(canonical: str, raw_heading: str, content: dict) -> None:
        if not content or not (content.get("blocks") or content.get("text")):
            return
        if canonical == "OTHERS":
            others.append({"heading": raw_heading, **content})
            return
        existing = out.get(canonical)
        if not existing:
            out[canonical] = content
        else:
            existing["blocks"] = (existing.get("blocks") or []) + (content.get("blocks") or [])
            existing["text"] = _sage_blocks_to_text(existing["blocks"])

    # TITLE from page <title> or dc.Title meta
    title_meta = soup.find("meta", attrs={"name": "dc.Title"})
    if title_meta and title_meta.get("content"):
        out["TITLE"] = _sage_clean(title_meta["content"])
    else:
        page_title = soup.find("title")
        if page_title:
            # Strip " - AuthorName, Year" suffix from SAGE title tags
            raw = _sage_clean(page_title.get_text(" ", strip=True))
            raw = re.sub(r"\s*-\s*.+?,\s*\d{4}\s*$", "", raw).strip()
            if raw:
                out["TITLE"] = raw

    # ABSTRACT from section#abstract
    abs_sec = soup.find("section", id="abstract")
    if abs_sec:
        h = abs_sec.find(["h2", "h3"])
        if h:
            h.decompose()
        blocks = _sage_collect_blocks(abs_sec) or []
        if not blocks:
            text = _sage_clean(abs_sec.get_text(" ", strip=True))
            if text:
                blocks = [{"type": "p", "text": text}]
        if blocks:
            out["ABSTRACT"] = _sage_build_content(blocks)

    # Named standalone sections (funding, acknowledgments, conflict, etc.)
    for sec_id, canonical in _SAGE_SECTION_ID_MAP.items():
        if canonical is None or sec_id == "abstract":
            continue
        sec = soup.find("section", id=sec_id)
        if not sec:
            continue
        blocks = _sage_collect_blocks(sec)
        if blocks:
            _merge(canonical, sec_id, _sage_build_content(blocks))

    # Body sections: top-level sec-X inside bodymatter.
    # Sections live inside a wrapper <div>, not as direct children, so use recursive=True
    # but skip sub-sections whose immediate parent is already a sec- section.
    bm = soup.find("section", id="bodymatter")
    if bm:
        for sec in bm.find_all("section", id=re.compile(r"^sec-")):
            if sec.find_parent("section", id=re.compile(r"^sec-")):
                continue  # skip nested sub-sections
            h_tag = sec.find(["h2", "h3", "h4"])
            if not h_tag:
                continue
            raw_heading = _sage_clean(h_tag.get_text(" ", strip=True))
            canonical = _sage_normalize_heading(raw_heading)
            blocks = _sage_collect_blocks(sec)
            if blocks:
                _merge(canonical, raw_heading, _sage_build_content(blocks))

    # KEYWORDS section (if not already from meta)
    if "KEYWORDS" not in out and keywords:
        out["KEYWORDS"] = {"blocks": [{"type": "list", "ordered": False, "items": keywords}],
                           "text": ", ".join(keywords)}

    if others:
        out["OTHERS"] = others

    return {"sections": out, "access_type": access_type, "keywords": keywords}


# ---------------------------------------------------------------------------
# Article data save locations (when running skipped flow):
#   - Issue URL articles (full metadata + authors): C:/sage_offline_uploads/offline_uploads/*.json
#   - Article URL data (jid, url, article_data for API): C:/sage_article_data_offline_uploads/offline_uploads/*.json
# ---------------------------------------------------------------------------


def remove_url_from_skipped_files(skipped_files, url):
    """Remove the line containing this URL from each skipped file (JSONL)."""
    url = (url or "").strip()
    if not url:
        return
    for path in skipped_files:
        path = os.path.abspath(path) if isinstance(path, str) else path
        if not os.path.exists(path):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            new_lines = []
            removed = False
            for line in lines:
                s = line.strip()
                if not s:
                    continue
                try:
                    obj = json.loads(s)
                    if (obj.get("url") or "").strip() == url:
                        removed = True
                        continue
                except json.JSONDecodeError:
                    pass
                new_lines.append(line if line.endswith("\n") else line + "\n")
            if removed:
                with open(path, "w", encoding="utf-8") as f:
                    f.writelines(new_lines)
                print(f"  → Removed URL from {os.path.basename(path)}")
        except Exception as e:
            print(f"  ⚠ Could not update {path}: {e}")


def _worker_safe_get(SKIPPED_FILE, driver, url, journal_id, retries=3, wait_time=180):
    """
    Same as safe_get but NEVER creates a new driver. Uses only the passed-in driver.
    Returns driver on success, or None if driver is dead or all retries fail. Use in workers so each worker keeps exactly one driver.
    """
    if driver is None:
        return None
    try:
        _ = driver.session_id
    except Exception:
        return None
    for attempt in range(1, retries + 1):
        try:
            driver.set_page_load_timeout(wait_time + 60)
            driver.get(url)
            WebDriverWait(driver, wait_time).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            src = driver.page_source.lower()
            if "ip has been blocked" in src or "ip blocked" in src:
                save_skipped(SKIPPED_FILE, journal_id, url, "IP blocked")
                return None
            start = time.time()
            while time.time() - start < wait_time:
                src = driver.page_source.lower()
                if "just a moment" in src or "verify you are human" in src or "checking your browser" in src:
                    time.sleep(5)
                else:
                    return driver
            raise TimeoutException("Cloudflare not cleared")
        except Exception as e:
            msg = str(e).lower()
            is_dead = (
                "connection refused" in msg or "max retries exceeded" in msg
                or "failed to establish a new connection" in msg or "invalid session" in msg
            )
            if is_dead:
                return None
            time.sleep(5)
    save_skipped(SKIPPED_FILE, journal_id, url, "Final failure")
    return driver


def remove_urls_from_skipped_files(skipped_files, processed_urls_set):
    """Remove all lines whose url is in processed_urls_set from each skipped file (one read/write per file)."""
    if not processed_urls_set:
        return
    for path in skipped_files:
        path = os.path.abspath(path) if isinstance(path, str) else path
        if not os.path.exists(path):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                lines = f.readlines()
            new_lines = []
            removed_count = 0
            for line in lines:
                s = line.strip()
                if not s:
                    continue
                try:
                    obj = json.loads(s)
                    if (obj.get("url") or "").strip() in processed_urls_set:
                        removed_count += 1
                        continue
                except json.JSONDecodeError:
                    pass
                new_lines.append(line if line.endswith("\n") else line + "\n")
            if removed_count:
                with open(path, "w", encoding="utf-8") as f:
                    f.writelines(new_lines)
                print(f"  → Removed {removed_count} URL(s) from {os.path.basename(path)}")
        except Exception as e:
            print(f"  ⚠ Could not update {path}: {e}")


def _process_skipped_chunk(args):
    """
    Worker for multiprocessing: process a chunk of entries with one driver.
    Returns list of URLs that were successfully processed (for removal from skipped files).
    Must be top-level for pickling. Uses force_close_driver so Chrome/chromedriver are always killed.
    worker_index is used to stagger driver creation so multiple workers don't race on the same chromedriver path.
    """
    import atexit
    (chunk, SKIPPED_FILE, database, base_url, article_urls_api_tpl, process_issue_urls, process_article_urls, worker_index) = args
    # Stagger driver creation: avoid all workers hitting undetected_chromedriver path at once (WinError 183)
    time.sleep(worker_index * 4)
    processed = []
    driver = None
    try:
        driver = create_driver(head)
        # Ensure driver is closed when this process exits (e.g. crash or Pool terminate)
        atexit.register(lambda: force_close_driver(driver) if driver else None)
    except Exception as ex:
        print(f"[worker] Failed to create driver: {ex}")
        return processed
    article_urls_cache = {}
    article_data_db = f"{database}_article_data"
    article_data_batch = []
    ARTICLE_DATA_BATCH_SIZE = 10
    try:
        for e in chunk:
            jid, url, reason = e["jid"], e["url"], e["reason"]
            if is_issue_url(url) and process_issue_urls and parse_issue_url and merge_issue_into_volume_backup:
                parsed = parse_issue_url(url)
                if not parsed:
                    continue
                volume = parsed.get("volume")
                issue_num = parsed.get("issue")
                year = parsed.get("year") or ""
                merge_issue_into_volume_backup(database, jid, url, volume, issue_num, year=year)
                if jid not in article_urls_cache:
                    try:
                        r = requests.get(
                            article_urls_api_tpl.format(database=database, jid=jid), timeout=60
                        )
                        data = r.json() if r.ok else {}
                        article_urls_cache[jid] = [
                            item.get("article_url")
                            for item in data.get("data", [])
                            if item.get("article_url")
                        ]
                    except Exception:
                        article_urls_cache[jid] = []
                article_urls = set(article_urls_cache.get(jid, []))
                batch = process_one_issue_url_sage(
                    SKIPPED_FILE, driver, jid, url, volume, issue_num,
                    published_year=year, article_urls=article_urls,
                    base_url=base_url, database=database,
                    safe_get_func=_worker_safe_get,
                )
                if batch:
                    save_offline(database, batch)
                processed.append(url)
            elif not is_issue_url(url) and process_article_urls:
                driver = _worker_safe_get(SKIPPED_FILE, driver, url, journal_id=jid)
                if driver is None:
                    break
                html = driver.page_source
                ad = extract_article_data(html) or []
                if ad:
                    article_data_batch.append({"jid": jid, "url": url, "article_data": ad})
                    if len(article_data_batch) >= ARTICLE_DATA_BATCH_SIZE:
                        save_offline(article_data_db, article_data_batch)
                        article_data_batch.clear()
                    processed.append(url)
            time.sleep(0.5)
        if article_data_batch:
            save_offline(article_data_db, article_data_batch)
    finally:
        if driver is not None:
            force_close_driver(driver)
    return processed


def collect_all_skipped_to_file(skipped_files, output_path="sage_all_skipped.txt", dedupe=True):
    """Load all URLs from skipped file(s), dedupe, write to one file. Returns list of entries."""
    if isinstance(skipped_files, str):
        skipped_files = [skipped_files]
    entries = load_all_skipped_urls(
        skipped_files, dedupe=dedupe, include_article_urls=True, include_issue_urls=True
    )
    if not entries:
        print("[*] No entries to write.")
        return []
    out_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        for e in entries:
            f.write(json.dumps({"jid": e["jid"], "url": e["url"], "reason": e["reason"]}) + "\n")
    print(f"[✔] Wrote {len(entries)} unique URLs → {out_path}")
    return entries


def load_article_urls_from_common_emails_json(json_path: str) -> list[dict]:
    """
    Load SAGE common_emails JSON produced by mando_db/update_databases.py (keyed by article_url).

    Example:
      {
        "database": "SAGE",
        "generated_at": "...",
        "data": {
          "https://.../doi/abs/...": {"article_id": "...", ...},
          ...
        }
      }
    Returns list of {"article_url": str, "article_id": str}.
    """
    json_path = os.path.abspath((json_path or "").strip())
    if not json_path or not os.path.exists(json_path):
        print(f"[common_emails_json] JSON not found: {json_path}")
        return []
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            root = json.load(f) or {}
    except Exception as e:
        print(f"[common_emails_json] Failed to read JSON: {e}")
        return []
    data = root.get("data") or {}
    out = []
    seen = set()
    for url, block in data.items():
        u = str(url or "").strip()
        if not u or u in seen:
            continue
        seen.add(u)
        aid = ""
        if isinstance(block, dict):
            aid = str(block.get("article_id") or "").strip()
        out.append({"article_url": u, "article_id": aid})
    print(f"[common_emails_json] Loaded {len(out)} unique URLs from {json_path}")
    return out


def _process_common_emails_chunk(args):
    """
    Worker: process a chunk of article URLs with one driver.
    Saves extracted payload to offline_uploads in batches.
    """
    import atexit

    (chunk, skipped_file, database, worker_index, shared_done, shared_total, lock, print_details) = args
    time.sleep(worker_index * 4)  # stagger driver creation
    driver = None
    processed = 0
    batch = []
    BATCH = 10
    out_db = f"{database}_common_emails_article_data"
    try:
        driver = create_driver(head)
        atexit.register(lambda: force_close_driver(driver) if driver else None)
        for rec in chunk:
            url = (rec.get("article_url") or "").strip()
            if not url:
                continue
            driver = _worker_safe_get(skipped_file, driver, url, journal_id=DATABASE)
            if driver is None:
                break
            html = driver.page_source
            ad = extract_article_data(html) or []
            if print_details:
                # Print per-URL extraction summary: author -> all emails
                mapped = []
                for au in ad:
                    if not isinstance(au, dict):
                        continue
                    nm = (au.get("author_name") or "").strip()
                    ems = au.get("emails") or ([au.get("email")] if au.get("email") else [])
                    ems = [str(e).strip() for e in ems if str(e).strip()]
                    if not ems:
                        continue
                    mapped.append((nm, ems, (au.get("author_type") or "").strip()))
                print(f"\n[w{worker_index}] URL: {url}\n  authors_with_emails={len(mapped)}", flush=True)
                for nm, ems, at in mapped[:25]:
                    print(f"  - {nm} [{at}] -> {', '.join(ems)}", flush=True)
                if len(mapped) > 25:
                    print(f"  ... and {len(mapped) - 25} more authors", flush=True)
            batch.append(
                {
                    "article_url": url,
                    "article_id": (rec.get("article_id") or "").strip(),
                    "article_data": ad,
                }
            )
            processed += 1
            if lock is not None and shared_done is not None:
                try:
                    with lock:
                        shared_done.value += 1
                        done_now = int(shared_done.value)
                        total_now = int(shared_total.value) if shared_total is not None else 0
                        # Single-line progress for the whole run
                        if total_now:
                            print(f"\r[common_emails] {done_now}/{total_now}", end="", flush=True)
                except Exception:
                    pass
            if len(batch) >= BATCH:
                save_offline(out_db, batch)
                batch.clear()
            time.sleep(0.5)
        if batch:
            save_offline(out_db, batch)
            batch.clear()
    finally:
        if driver is not None:
            force_close_driver(driver)
    return processed


def run_common_emails_json_flow(json_path: str, workers: int = 5, print_details: bool = False):
    """
    Non-skipped flow:
    - Read all URLs from common_emails JSON
    - Process with N browser workers (default 5)
    - Save extracted data to C:/<db>_common_emails_article_data_offline_uploads/offline_uploads/*.json
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    skipped_file = os.path.join(script_dir, "sage_common_emails_skipped.txt")
    entries = load_article_urls_from_common_emails_json(json_path)
    if not entries:
        return
    total = len(entries)
    workers = max(1, int(workers or 1))
    if workers <= 1:
        # sequential (no shared counter needed)
        _process_common_emails_chunk((entries, skipped_file, DATABASE, 0, None, None, None, print_details))
        return
    chunks = [[] for _ in range(workers)]
    for i, rec in enumerate(entries):
        chunks[i % workers].append(rec)
    manager = multiprocessing.Manager()
    shared_done = manager.Value("i", 0)
    shared_total = manager.Value("i", total)
    lock = manager.Lock()

    args_list = [
        (chunk, skipped_file, DATABASE, worker_index, shared_done, shared_total, lock, print_details)
        for worker_index, chunk in enumerate(chunks)
    ]
    print(f"[*] Processing {len(entries)} common_emails URLs with {workers} workers... (progress 0/{total})")
    try:
        from driver_state import kill_orphan_drivers
        kill_orphan_drivers()
    except Exception:
        pass
    with multiprocessing.Pool(processes=workers) as pool:
        done_counts = pool.map(_process_common_emails_chunk, args_list)
    print()
    print(f"[*] common_emails JSON flow finished. processed={sum(done_counts)}/{total}")


def process_one_issue_url_sage(
    SKIPPED_FILE,
    driver,
    journal_id,
    issue_url,
    volume,
    issue_num,
    published_year,
    article_urls,
    base_url="https://journals.sagepub.com/",
    database=DATABASE,
    safe_get_func=None,
):
    """Load one SAGE issue page, extract articles, fetch each and build combined_batch. Returns list. safe_get_func: use this instead of safe_get (e.g. _worker_safe_get to keep one driver per worker)."""
    get_url = safe_get_func or safe_get
    driver = get_url(SKIPPED_FILE, driver, issue_url, journal_id=journal_id)
    if not driver:
        return []
    time.sleep(3)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    sections = soup.find_all("div", class_="issue-item__container")
    if not sections:
        return []
    combined_batch = []
    seen_in_issue = set()
    for container in sections:
        title_tags = container.find_all("div", class_="issue-item__title") if container else []
        for title_tag in title_tags:
            article_url = ""
            title = ""
            a = title_tag.find("a", href=True)
            if not a:
                continue
            article_url = urljoin(base_url, a["href"])
            title = title_tag.a.get_text(strip=True) if title_tag.a else ""
            if not article_url or article_url in article_urls:
                continue
            if article_url in seen_in_issue:
                continue
            seen_in_issue.add(article_url)
            published_date = ""
            article_type = ""
            header = container.find("div", class_="issue-item__header")
            if header:
                spans = header.find_all("span")
                if len(spans) >= 2:
                    article_type = spans[1].get_text(strip=True)
                for sp in spans:
                    text = sp.get_text(strip=True)
                    if "First published" in text:
                        published_date = text.replace("First published", "").strip()
                        break
            pdf_url = ""
            actions = container.find("div", class_="issue-item__actions")
            if actions:
                a_pdf = actions.find("a", title="download", href=True)
                if a_pdf:
                    pdf_url = urljoin(base_url, a_pdf["href"])
            driver = get_url(SKIPPED_FILE, driver, article_url, journal_id=journal_id)
            if not driver:
                continue
            time.sleep(2)
            soup_article = BeautifulSoup(driver.page_source, "html.parser")
            doi_tag = soup_article.find("div", class_="doi")
            doi = doi_tag.get_text(separator=" ", strip=True) if doi_tag else ""
            abstract = ""
            abstract_content = soup_article.find("section", id="abstract")
            if abstract_content:
                h2_tag = abstract_content.find("h2")
                if h2_tag:
                    h2_tag.decompose()
                abstract = abstract_content.get_text(separator=" ", strip=True)
            acknowledgment = ""
            acknowledgment_tag = soup_article.find("section", id="acknowledgments")
            if acknowledgment_tag:
                h2_tag = acknowledgment_tag.find("h2")
                if h2_tag:
                    h2_tag.decompose()
                acknowledgment = acknowledgment_tag.get_text(strip=True)
            funding = ""
            funding_tag = soup_article.find("section", id="funding")
            if funding_tag:
                h2_tag = funding_tag.find("h2")
                if h2_tag:
                    h2_tag.decompose()
                funding = funding_tag.get_text(strip=True) if funding_tag else ""
            _page_html = driver.page_source
            global _CURRENT_BROWSER
            _CURRENT_BROWSER = driver
            try:
                _extras = extract_sections_and_access(_page_html)
            finally:
                _CURRENT_BROWSER = None
            combined_batch.append({
                "article_link": {
                    "jid": journal_id,
                    "article_title": title,
                    "article_type": article_type,
                    "article_url": article_url,
                    "abstract": abstract,
                    "pdf": pdf_url,
                    "doi": doi,
                    "published_year": published_year or "",
                    "published_date": published_date,
                    "volume": str(volume),
                    "issue": str(issue_num),
                    "funding": funding,
                    "funding_information": [],
                    "acknowledgement": acknowledgment,
                    "access_type": _extras["access_type"],
                    "sections": _extras["sections"],
                    "keywords": _extras["keywords"],
                    "html": _page_html,
                    "success": True,
                    "queue": True,
                },
                "article_data": extract_article_data(_page_html) or [],
            })
            # Debug: same as OUP / spring_nature
            article_data_list = combined_batch[-1].get("article_data") or []
            print("\n" + "=" * 60)
            print("EXTRACTED VALUES (debug)")
            print("=" * 60)
            print(f"  jid                 : {journal_id}")
            print(f"  article_url         : {article_url}")
            print(f"  article_title       : {title}")
            print(f"  article_type        : {article_type}")
            print(f"  published_year      : {published_year or ''}  |  published_date: {published_date}")
            print(f"  volume              : {volume}  |  issue: {issue_num}")
            print(f"  doi                 : {doi}")
            print(f"  pdf                 : {pdf_url}")
            print(f"  abstract            : {(abstract or '')[:100]}...")
            print(f"  funding             : {(funding or '')[:80]}...")
            print(f"  acknowledgement     : {(acknowledgment or '')[:80]}...")
            print(f"  article_data (count): {len(article_data_list)} authors")
            for i, ad in enumerate(article_data_list[:5]):
                print(f"    author[{i}]       : {ad.get('author_name')} | {ad.get('author_type')} | {ad.get('country')}")
            if len(article_data_list) > 5:
                print(f"    ... and {len(article_data_list) - 5} more authors")
            print("=" * 60 + "\n")
            print(f"✅ Processed: {title}")
    return combined_batch


def collect_article_links_from_issue(driver, SKIPPED_FILE, journal_id, issue_url, volume, issue_num, published_year, article_urls, base_url="https://journals.sagepub.com/"):
    """
    Phase 1: Visit one issue page and collect article link metadata WITHOUT visiting each article page.
    Returns list of dicts: {jid, article_url, title, article_type, published_date, pdf, volume, issue, published_year}
    """
    driver = safe_get(SKIPPED_FILE, driver, issue_url, journal_id=journal_id)
    if not driver:
        return driver, []
    time.sleep(3)
    soup = BeautifulSoup(driver.page_source, "html.parser")
    sections = soup.find_all("div", class_="issue-item__container")
    links = []
    seen = set()
    for container in sections:
        for title_tag in container.find_all("div", class_="issue-item__title"):
            a = title_tag.find("a", href=True)
            if not a:
                continue
            article_url = urljoin(base_url, a["href"])
            if not article_url or article_url in article_urls or article_url in seen:
                continue
            seen.add(article_url)
            title = a.get_text(strip=True)
            article_type = ""
            published_date = ""
            header = container.find("div", class_="issue-item__header")
            if header:
                spans = header.find_all("span")
                if len(spans) >= 2:
                    article_type = spans[1].get_text(strip=True)
                for sp in spans:
                    t = sp.get_text(strip=True)
                    if "First published" in t:
                        published_date = t.replace("First published", "").strip()
                        break
            pdf_url = ""
            actions = container.find("div", class_="issue-item__actions")
            if actions:
                a_pdf = actions.find("a", title="download", href=True)
                if a_pdf:
                    pdf_url = urljoin(base_url, a_pdf["href"])
            links.append({
                "jid": journal_id,
                "article_url": article_url,
                "article_title": title,
                "article_type": article_type,
                "published_date": published_date,
                "pdf": pdf_url,
                "volume": str(volume),
                "issue": str(issue_num),
                "published_year": str(published_year or ""),
            })
            print(f"  [link] {title[:80]} → {article_url}")
    return driver, links


def enrich_article_links_with_data(
    DATA_STATE_FILE,
    SKIPPED_FILE,
    queue_file: str,
    driver=None,
    base_url="https://journals.sagepub.com/",
):
    """
    Phase 2: Read article links from queue_file (JSONL written by --links-only),
    visit each article page, extract doi/abstract/funding/ack/article_data,
    save via save_offline.

    Resume: DATA_STATE_FILE (e.g. sagepub_last1_data.json) stores the last processed URL.
    On restart, all queue entries up to and including last_url are skipped instantly
    (no browser needed) — then processing continues from the next entry.

    Saved data location: C:/sagepub_offline_uploads/offline_uploads/*.json
    Queue file:          sage/sagepub_links_queue_{part}.jsonl
    State file:          sage/sagepub_last{part}_data.json  {"last_url": "...", "total_done": N}
    """
    if not os.path.exists(queue_file):
        print(f"[data-only] Queue file not found: {queue_file}")
        return driver

    # Load last state — find where we stopped
    last_url = None
    total_done = 0
    if os.path.exists(DATA_STATE_FILE):
        try:
            with open(DATA_STATE_FILE, "r", encoding="utf-8") as f:
                st = json.load(f)
            last_url = st.get("last_url", "")
            total_done = int(st.get("total_done", 0))
            if last_url:
                print(f"[data-only] Resuming after: {last_url}  (done so far: {total_done})")
        except Exception:
            pass

    def _save_data_state(url, done_count):
        try:
            with open(DATA_STATE_FILE, "w", encoding="utf-8") as f:
                json.dump({"last_url": url, "total_done": done_count}, f, indent=2)
        except Exception as e:
            print(f"[data-only] Could not save state: {e}")

    if not driver:
        driver = create_driver(head)

    combined_batch = []
    skipping = bool(last_url)   # True until we pass last_url in the queue

    with open(queue_file, "r", encoding="utf-8") as f:
        for raw in f:
            raw = raw.strip()
            if not raw:
                continue
            try:
                rec = json.loads(raw)
            except Exception:
                continue

            article_url = rec.get("article_url", "")
            if not article_url:
                continue

            # Fast-skip everything up to and including last_url
            if skipping:
                if article_url == last_url:
                    skipping = False
                print(f"  [skip] {article_url}")
                continue

            journal_id = rec.get("jid", "")
            print(f"[data] {article_url}")

            driver = safe_get(SKIPPED_FILE, driver, article_url, journal_id=journal_id)
            if not driver:
                continue
            time.sleep(4)
            soup = BeautifulSoup(driver.page_source, "html.parser")

            doi = ""
            doi_tag = soup.find("div", class_="doi")
            if doi_tag:
                doi = doi_tag.get_text(separator=" ", strip=True)

            abstract = ""
            abstract_sec = soup.find("section", id="abstract")
            if abstract_sec:
                h2 = abstract_sec.find("h2")
                if h2:
                    h2.decompose()
                abstract = abstract_sec.get_text(separator=" ", strip=True)

            acknowledgment = ""
            ack_sec = soup.find("section", id="acknowledgments")
            if ack_sec:
                h2 = ack_sec.find("h2")
                if h2:
                    h2.decompose()
                acknowledgment = ack_sec.get_text(strip=True)

            funding = ""
            fund_sec = soup.find("section", id="funding")
            if fund_sec:
                h2 = fund_sec.find("h2")
                if h2:
                    h2.decompose()
                funding = fund_sec.get_text(strip=True)

            _page_html = driver.page_source
            global _CURRENT_BROWSER
            _CURRENT_BROWSER = driver
            try:
                _extras = extract_sections_and_access(_page_html)
            finally:
                _CURRENT_BROWSER = None
            article_data = extract_article_data(_page_html) or []

            combined_batch.append({
                "article_link": {
                    "jid": journal_id,
                    "article_title": rec.get("article_title", ""),
                    "article_type": rec.get("article_type", ""),
                    "article_url": article_url,
                    "abstract": abstract,
                    "pdf": rec.get("pdf", ""),
                    "doi": doi,
                    "published_year": rec.get("published_year", ""),
                    "published_date": rec.get("published_date", ""),
                    "volume": rec.get("volume", ""),
                    "issue": rec.get("issue", ""),
                    "funding": funding,
                    "funding_information": [],
                    "acknowledgement": acknowledgment,
                    "access_type": _extras["access_type"],
                    "sections": _extras["sections"],
                    "keywords": _extras["keywords"],
                    "html": _page_html,
                    "success": True,
                    "queue": False,
                },
                "article_data": article_data,
            })

            total_done += 1
            _save_data_state(article_url, total_done)
            print(f"  doi={doi[:40]}  authors={len(article_data)}  done={total_done}")

            if len(combined_batch) >= BATCH_SIZE:
                save_offline(DATABASE, combined_batch)
                combined_batch.clear()

    if combined_batch:
        save_offline(DATABASE, combined_batch)
        combined_batch.clear()

    print(f"[data-only] Done. Total enriched: {total_done}")
    return driver


def process_skipped_urls(
    skipped_files,
    SKIPPED_FILE,
    driver=None,
    database=DATABASE,
    base_url="https://journals.sagepub.com/",
    article_urls_api_tpl="http://139.84.134.18:8002/{database}/all_article_urls?jid={jid}",
    write_all_skipped_path=None,
    process_issue_urls=True,
    process_article_urls=True,
):
    """Load skipped URLs, dedupe. Issue URLs: merge backup + extract + save_offline. Article URLs: save to article_data folder every 10, remove from files."""
    if isinstance(skipped_files, str):
        skipped_files = [skipped_files]
    entries = load_all_skipped_urls(
        skipped_files, dedupe=True, include_article_urls=True, include_issue_urls=True
    )
    if not entries:
        print("[*] No skipped URLs to process.")
        return
    if write_all_skipped_path:
        out_path = os.path.abspath(write_all_skipped_path)
        os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            for e in entries:
                f.write(json.dumps({"jid": e["jid"], "url": e["url"], "reason": e["reason"]}) + "\n")
        print(f"[✔] Wrote {len(entries)} unique URLs → {out_path}")
    if not driver:
        driver = create_driver(head)
    article_urls_cache = {}
    article_data_db = f"{database}_article_data"
    article_data_batch = []
    ARTICLE_DATA_BATCH_SIZE = 10
    try:
        for i, e in enumerate(entries, start=1):
            jid, url, reason = e["jid"], e["url"], e["reason"]
            print(f"[{i}/{len(entries)}] {url[:80]}...")
            if is_issue_url(url) and process_issue_urls and parse_issue_url and merge_issue_into_volume_backup:
                parsed = parse_issue_url(url)
                if not parsed:
                    continue
                volume = parsed.get("volume")
                issue_num = parsed.get("issue")
                year = parsed.get("year") or ""
                merge_issue_into_volume_backup(database, jid, url, volume, issue_num, year=year)
                if jid not in article_urls_cache:
                    try:
                        r = requests.get(
                            article_urls_api_tpl.format(database=database, jid=jid), timeout=60
                        )
                        data = r.json() if r.ok else {}
                        article_urls_cache[jid] = [
                            item.get("article_url")
                            for item in data.get("data", [])
                            if item.get("article_url")
                        ]
                    except Exception:
                        article_urls_cache[jid] = []
                article_urls = set(article_urls_cache.get(jid, []))
                batch = process_one_issue_url_sage(
                    SKIPPED_FILE,
                    driver,
                    jid,
                    url,
                    volume,
                    issue_num,
                    published_year=year,
                    article_urls=article_urls,
                    base_url=base_url,
                    database=database,
                )
                if batch:
                    save_offline(database, batch)
                    print(f"  → Saved {len(batch)} articles from issue to offline_uploads")
                remove_url_from_skipped_files(skipped_files, url)
            elif not is_issue_url(url) and process_article_urls:
                driver = safe_get(SKIPPED_FILE, driver, url, journal_id=jid)
                if not driver:
                    time.sleep(1)
                    continue
                html = driver.page_source
                ad = extract_article_data(html) or []
                if ad:
                    article_data_batch.append({"jid": jid, "url": url, "article_data": ad})
                    if len(article_data_batch) >= ARTICLE_DATA_BATCH_SIZE:
                        save_offline(article_data_db, article_data_batch)
                        article_data_batch.clear()
                        print("  → Flushed article_data_by_url batch to offline_uploads")
                    remove_url_from_skipped_files(skipped_files, url)
            time.sleep(1)
        if article_data_batch:
            save_offline(article_data_db, article_data_batch)
            article_data_batch.clear()
            print("  → Flushed final article_data_by_url batch to offline_uploads")
    finally:
        if driver is not None:
            try:
                force_close_driver(driver)
            except Exception:
                pass


# Number of parallel workers for skipped flow (each has its own browser driver)
SKIPPED_FLOW_WORKERS = 5


def run_skipped_flow(workers=SKIPPED_FLOW_WORKERS):
    """
    Run skipped-URL flow: collect sage_skipped*.txt, dedupe, process in parallel, remove from files.
    Article data save locations:
      - Issue URL articles: C:/sage_offline_uploads/offline_uploads/*.json
      - Article URL data (by url): C:/sage_article_data_offline_uploads/offline_uploads/*.json
    """
    import multiprocessing
    script_dir = os.path.dirname(os.path.abspath(__file__))
    skipped_files = []
    for f in os.listdir(script_dir):
        if f.startswith("sage_skipped") and f.endswith(".txt") and "all_skipped" not in f and "retry" not in f:
            skipped_files.append(os.path.join(script_dir, f))
    for name in ["sage_skipped1.txt", "sage_skipped.txt"]:
        p = os.path.join(script_dir, name)
        if os.path.exists(p) and p not in skipped_files:
            skipped_files.append(p)
    if not skipped_files:
        print("[!] No sage_skipped*.txt files found in:", script_dir)
        return
    skipped_files = sorted(set(skipped_files))
    all_skipped_path = os.path.join(script_dir, "sage_all_skipped.txt")
    retry_skipped_log = os.path.join(script_dir, "sage_skipped_retry.txt")
    print(f"[*] Skipped files: {skipped_files}")
    print("[*] Article data will be saved to:")
    print("    - Issue articles: C:/sage_offline_uploads/offline_uploads/")
    print("    - Article-by-URL data: C:/sage_article_data_offline_uploads/offline_uploads/")
    collect_all_skipped_to_file(skipped_files, output_path=all_skipped_path, dedupe=True)
    entries = load_all_skipped_urls(
        skipped_files, dedupe=True, include_article_urls=True, include_issue_urls=True
    )
    if not entries:
        print("[*] No entries to process.")
        return
    workers = max(1, workers)
    if workers <= 1:
        process_skipped_urls(
            skipped_files,
            retry_skipped_log,
            database=DATABASE,
            write_all_skipped_path=None,
            process_issue_urls=True,
            process_article_urls=True,
        )
    else:
        # Always create exactly `workers` chunks so we get exactly `workers` drivers (one per worker)
        chunks = [[] for _ in range(workers)]
        for i, e in enumerate(entries):
            chunks[i % workers].append(e)
        article_urls_api_tpl = "http://139.84.134.18:8002/{database}/all_article_urls?jid={jid}"
        args_list = [
            (
                chunk,
                retry_skipped_log,
                DATABASE,
                "https://journals.sagepub.com/",
                article_urls_api_tpl,
                True,
                True,
                worker_index,
            )
            for worker_index, chunk in enumerate(chunks)
        ]
        print(f"[*] Processing {len(entries)} URLs with {workers} workers...")
        try:
            from driver_state import kill_orphan_drivers
            kill_orphan_drivers()
        except Exception:
            pass
        with multiprocessing.Pool(processes=workers) as pool:
            results = pool.map(_process_skipped_chunk, args_list)
        processed_set = set()
        for r in results:
            processed_set.update(r)
        if processed_set:
            remove_urls_from_skipped_files(skipped_files, processed_set)
            print(f"[*] Removed {len(processed_set)} processed URLs from skipped files.")
    print("[*] Skipped flow finished.")


def _normalize_journal_url(u: str) -> str:
    u = (u or "").strip()
    if not u:
        return ""
    # Normalize scheme + casing + trailing slash for reliable comparisons
    u = u.replace("http://", "https://")
    u = u.rstrip("/")
    return u.lower()


def _fetch_existing_journal_urls_from_api(database: str) -> set:
    """Fetch all existing journals from API_ROOT/{database}/journals and return normalized journal_url set."""
    existing = set()
    page = 1
    while True:
        try:
            resp = requests.get(f"{API_ROOT}/{database}/journals?page={page}", timeout=60)
            if resp.status_code != 200:
                break
            data = resp.json() if resp.content else {}
            records = data.get("records", []) or []
            if not records:
                break
            for r in records:
                url = r.get("journal_url") or ""
                n = _normalize_journal_url(url)
                if n:
                    existing.add(n)
            page += 1
        except Exception as e:
            print(f"⚠️ API fetch error (page {page}): {e}")
            break
    return existing


def _compare_and_push_new(journals: list, database: str):
    """Compare given journals with DB by journal_url only and push only new ones."""
    print("\n[DB] Fetching existing journals from API for comparison...")
    existing_urls = _fetch_existing_journal_urls_from_api(database)
    scraped_unique = set()
    new_unique = []
    new_seen = set()
    for j in journals:
        nurl = _normalize_journal_url(j.get("journal_url", ""))
        if not nurl:
            continue
        scraped_unique.add(nurl)
        if nurl in existing_urls:
            continue
        if nurl in new_seen:
            continue
        new_seen.add(nurl)
        new_unique.append(j)

    print(f"[DB] Old journals (unique by journal_url): {len(existing_urls)}")
    print(f"[DB] Scraped journals (unique by journal_url): {len(scraped_unique)}")
    print(f"[DB] New journals to push (unique by journal_url): {len(new_unique)}")

    if new_unique:
        try:
            post_journals(new_unique, database)
            print(f"[DB] Pushed new journals: {len(new_unique)}")
        except Exception as e:
            print(f"[DB] ❌ Push failed: {e}")
    else:
        print("[DB] No new journals to push.")


def _load_journals_from_excel(path: str) -> list:
    """Load journals from an existing Excel file. Supports both 'old' and 'new' header formats."""
    print(f"[EXCEL] Loading journals from {path}")
    wb = load_workbook(path)
    ws = wb.active
    journals = []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return journals
    header = [str(h or "").strip().lower() for h in rows[0]]

    def col_idx(name_variants):
        for variant in name_variants:
            if variant in header:
                return header.index(variant)
        return None

    idx_publisher = col_idx(["publisher"])
    idx_name = col_idx(["journal_name", "name"])
    idx_jurl = col_idx(["journal_url"])
    idx_aurl = col_idx(["archive_page_url", "archive_url"])
    idx_issn = col_idx(["issn", "eissn"])
    idx_pssn = col_idx(["pssn", "print_issn"])
    idx_remarks = col_idx(["remarks", "remark"])

    def _cell(row, idx):
        if idx is None:
            return ""
        v = row[idx]
        if v is None:
            return ""
        # openpyxl returns typed values (int, float, datetime) — always stringify
        return str(v).strip()

    for row in rows[1:]:
        if not any(row):
            continue
        j = {
            "publisher": _cell(row, idx_publisher),
            "journal_name": _cell(row, idx_name),
            "journal_url": _cell(row, idx_jurl),
            "archive_page_url": _cell(row, idx_aurl),
            "issn": _cell(row, idx_issn),
            "pssn": _cell(row, idx_pssn),
            "remarks": _cell(row, idx_remarks),
        }
        journals.append(j)
    print(f"[EXCEL] Loaded {len(journals)} rows from Excel")
    return journals


def push_excel_to_db(excel_path: str, database: str = DATABASE, push_all: bool = False, batch_size: int = 200):
    """
    Push journals from an Excel into the DB in batches to avoid 500 errors on large payloads.
    - push_all=False: compare by journal_url and push only new ones (safe default)
    - push_all=True : push everything in batches of batch_size (may create duplicates)
    """
    journals = _load_journals_from_excel(excel_path)
    if not journals:
        print("[EXCEL] No rows to push.")
        return
    if push_all:
        print(f"[DB] Pushing ALL {len(journals)} rows in batches of {batch_size}...")
        pushed = 0
        for start in range(0, len(journals), batch_size):
            chunk = journals[start:start + batch_size]
            post_journals(chunk, database)
            pushed += len(chunk)
            print(f"[DB] Progress: {pushed}/{len(journals)}")
        print(f"[DB] Done. Total attempted: {len(journals)}")
    else:
        _compare_and_push_new(journals, database)


def start_process(url, base_url, push_new_to_db=False):
    driver = None
    try:
        driver = create_driver(False)
        all_journals = []
        previous_signature = None  # stable signature of extracted journals on previous page
        consecutive_empty_pages = 0  # stop when we extract 0 journals twice in a row
        page = 0
        use_template = "{pagination}" in url
        MAX_PAGES_SAFE = 500  # hard stop to prevent infinite loops
        seen_effective_urls = set()  # stop if site redirects to same page repeatedly
        EXPECTED_TOTAL = 1505
        EXPECTED_PER_PAGE = 24
        LAST_PAGE_NUMBER = 63  # up to page 63; last page may have <24
        url_occurrence = {}  # journal_url -> 1,2,3... (do not dedupe; track repeats)

        while True:
            if page >= MAX_PAGES_SAFE:
                print(f"   (hit MAX_PAGES_SAFE={MAX_PAGES_SAFE}, stopping)")
                break
            page_url = url.format(pagination=page) if use_template else f"{url}{page}"
            print(f"🌀 Processing Page {page + 1} → {page_url}")
            driver.get(page_url)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            # Give page time to fully render results (24 per page).
            time.sleep(2)
            try:
                WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "article.product-search-solr, div.item__body.clearfix"))
                )
            except TimeoutException:
                pass
            time.sleep(10)
            # If count is less than 24, keep waiting (check every 5s) up to 180s total,
            # but only for pages before the last page (page 63 can legitimately have <24).
            start_wait = time.time()
            count = 0
            while True:
                html_tmp = driver.page_source
                soup_tmp = BeautifulSoup(html_tmp, "html.parser")
                if soup_tmp.find_all("div", class_="item__body clearfix"):
                    count = len(soup_tmp.find_all("div", class_="item__body clearfix"))
                else:
                    count = len(soup_tmp.find_all("article", class_=re.compile(r"\bproduct-search-solr\b", re.I)))

                elapsed = time.time() - start_wait
                # For last page (63) just use whatever is there; no long wait.
                if page + 1 >= LAST_PAGE_NUMBER:
                    break
                if count >= EXPECTED_PER_PAGE:
                    break
                if elapsed >= 180:
                    print(f"   ⚠ Waited 180s but still only {count}/{EXPECTED_PER_PAGE} items; continuing anyway")
                    break
                print(f"   … only {count}/{EXPECTED_PER_PAGE} items loaded; waiting 5s (elapsed {int(elapsed)}s)")
                time.sleep(5)

            effective_url = (getattr(driver, "current_url", None) or "").strip()
            if effective_url:
                if effective_url in seen_effective_urls:
                    print(f"   (redirected to already-seen page → stop) {effective_url}")
                    break
                seen_effective_urls.add(effective_url)
            current_page_html = driver.page_source
            soup = BeautifulSoup(current_page_html, "html.parser")

            extracted_this_page = []

            # --- Extract mode A: journals.sagepub.com listing cards (old flow) ---
            items = soup.find_all("div", class_="item__body clearfix")
            if items:
                for item in items:
                    tag = item.find("h3", class_="heading-search item__title")
                    if not tag:
                        continue
                    a_tag = tag.find("a", href=True)
                    if not a_tag:
                        continue
                    name_tag = a_tag.find("span", class_="hlFld-Title")
                    journal_name = name_tag.get_text(strip=True) if name_tag else a_tag.get_text(strip=True)
                    href = a_tag.get("href", "") or ""
                    if not href:
                        continue
                    journal_url = urljoin(base_url, href)
                    archive_url = journal_url.replace("/home/", "/loi/")

                    pssn_label = item.find("span", class_="meta__label meta__issns")
                    eissn_label = item.find("span", class_="meta__label meta__eissn")
                    pssn = pssn_label.find_next("span").get_text(strip=True) if pssn_label else ""
                    issn = eissn_label.find_next("span").get_text(strip=True) if eissn_label else ""

                    extracted_this_page.append({
                        "publisher": "SAGE",
                        "journal_name": journal_name,
                        "journal_url": journal_url,
                        "archive_page_url": archive_url,
                        "issn": issn,
                        "pssn": pssn,
                    })
            else:
                # --- Extract mode B: uk.sagepub.com product results ---
                # Pattern from `sage_journals_ref.html`:
                # <article class="product-search-solr" node-type="journal"> ... ISSN: 24557471 ... <a alt="Read Online" href="http://journals.sagepub.com/home/urb">Read Online</a>
                articles = soup.find_all("article", class_=re.compile(r"\bproduct-search-solr\b", re.I))
                if not articles:
                    # fallback: any <article node-type="journal">
                    articles = soup.find_all("article", attrs={"node-type": "journal"})

                for art in articles:
                    # Only journals
                    if (art.get("node-type") or "").lower() not in ("", "journal"):
                        continue

                    # Title link (on uk.sagepub.com)
                    title = ""
                    title_a = art.find("a", href=True, class_=re.compile(r"anchor-camouflaged", re.I))
                    if not title_a:
                        # fallback: first internal link to product page
                        title_a = art.find("a", href=re.compile(r"^/en-gb/afr/", re.I))
                    if title_a:
                        title = (title_a.get_text(strip=True) or "").strip()

                    # ISSN appears like: "ISSN:\n    24557471"
                    art_text = art.get_text(" ", strip=True)
                    pssn = ""
                    m = re.search(r"\bISSN:\s*([0-9Xx-]{4,})\b", art_text, flags=re.I)
                    if m:
                        pssn = m.group(1).strip()

                    # Read Online link: usually absolute to journals.sagepub.com/home/xxx
                    read_a = art.find("a", href=True, attrs={"alt": re.compile(r"read online", re.I)})
                    if not read_a:
                        read_a = art.find("a", href=True, string=re.compile(r"^\s*Read Online\s*$", re.I))
                    read_href = (read_a.get("href") or "").strip() if read_a else ""
                    if not read_href:
                        continue
                    journal_url = urljoin(base_url, read_href)
                    archive_url = journal_url.replace("/home/", "/loi/")

                    extracted_this_page.append({
                        "publisher": "SAGE",
                        "journal_name": title or "",
                        "journal_url": journal_url,
                        "archive_page_url": archive_url,
                        "issn": "",   # uk.sagepub.com page shows one ISSN; store in pssn for now
                        "pssn": pssn,
                    })

            # Stop condition: no journals extracted (not just "no divs found")
            if not extracted_this_page:
                consecutive_empty_pages += 1
                print(f"   (0 journals extracted on page, empty count {consecutive_empty_pages})")
                if consecutive_empty_pages >= 2:
                    print("   (0 journals extracted second time → stop)")
                    break
                page += 1
                continue
            consecutive_empty_pages = 0

            # Stable signature: ordered list of journal_url values
            current_signature = [j.get("journal_url", "") for j in extracted_this_page if j.get("journal_url")]
            if previous_signature is not None and current_signature == previous_signature:
                # keep duplicates from this repeated page, then stop (so you can see occurrence=2)
                print("   (same journal list as previous page → repeated page; collecting once then stop)")
                repeated_page = True
            else:
                repeated_page = False
            previous_signature = current_signature

            if len(extracted_this_page) != EXPECTED_PER_PAGE:
                print(f"   ⚠ Collected {len(extracted_this_page)} journals on page {page + 1} (expected {EXPECTED_PER_PAGE})")
            else:
                print(f"   Collected {len(extracted_this_page)} journals on page {page + 1}")
            for idx, j in enumerate(extracted_this_page, start=1):
                name = (j.get("journal_name") or "").strip() or "(no title found)"
                pssn = (j.get("pssn") or "").strip()
                eissn = (j.get("issn") or "").strip()
                jurl = (j.get("journal_url") or "").strip()
                url_occurrence[jurl] = url_occurrence.get(jurl, 0) + 1
                occ = url_occurrence[jurl]
                j["page_no"] = page + 1
                j["item_no"] = idx
                j["occurrence"] = occ
                print(f"✅ P{page + 1:02d} #{idx:02d} | occ:{occ} | {name} | PSSN:{pssn} | EISSN:{eissn} | {jurl}")
                all_journals.append(j)

            page += 1
            if repeated_page:
                break

        print(f"\n✅ Total Journals Extracted: {len(all_journals)} (expected {EXPECTED_TOTAL}, {EXPECTED_PER_PAGE}/page; stopped after page {page})")
        # Save "all journals" to Excel for checking
        _script_dir = os.path.dirname(os.path.abspath(__file__))
        output_excel = os.path.join(_script_dir, f"{DATABASE}_journals_check_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "SAGE Journals"
        headers = ["row_no", "page_no", "item_no", "occurrence", "publisher", "journal_name", "journal_url", "archive_page_url", "issn", "pssn"]
        ws.append(headers)
        for row_no, j in enumerate(all_journals, start=1):
            ws.append([
                row_no,
                j.get("page_no", ""),
                j.get("item_no", ""),
                j.get("occurrence", ""),
                j.get("publisher", ""),
                j.get("journal_name", ""),
                j.get("journal_url", ""),
                j.get("archive_page_url", ""),
                j.get("issn", ""),
                j.get("pssn", ""),
            ])
        wb.save(output_excel)
        print(f"Saved: {output_excel}")

        # Compare with DB by journal_url only and push only the new journals
        if push_new_to_db:
            _compare_and_push_new(all_journals, DATABASE)

        return all_journals
    finally:
        if driver is not None:
            try:
                force_close_driver(driver)
            except Exception:
                pass

def _extract_show_publications_journals_from_soup(
    soup: BeautifulSoup,
    base_url: str = "https://journals.sagepub.com",
    include_books: bool = False,
):
    """
    Extract journals from `https://journals.sagepub.com/action/showPublications` HTML.
    By default, skips book entries (li.book__item) so output is journals only.
    """
    extracted = []
    results_ul = soup.find("ul", class_=re.compile(r"\bsearch-result__publications\b"))
    if not results_ul:
        return extracted

    for li in results_ul.find_all("li", class_=re.compile(r"\bsearch__item\b")):
        li_cls = " ".join(li.get("class") or [])
        is_book = "book__item" in li_cls
        if is_book and not include_books:
            continue

        body = li.find("div", class_=re.compile(r"\bitem__body\b"))
        if not body:
            continue

        title_h3 = body.find("h3", class_=re.compile(r"\bitem__title\b"))
        if not title_h3:
            continue

        a = title_h3.find("a", href=True)
        if not a:
            continue

        name_tag = a.find("span", class_="hlFld-Title")
        journal_name = (name_tag.get_text(strip=True) if name_tag else a.get_text(strip=True)).strip()
        href = (a.get("href") or "").strip()
        if not href:
            continue

        journal_url = urljoin(base_url, href)
        archive_url = journal_url.replace("/home/", "/loi/")

        pssn = ""
        eissn = ""
        pssn_label = body.find("span", class_=re.compile(r"\bmeta__issns\b"))
        if pssn_label:
            pssn_span = pssn_label.find_next("span")
            pssn = pssn_span.get_text(strip=True) if pssn_span else ""
        eissn_label = body.find("span", class_=re.compile(r"\bmeta__eissn\b"))
        if eissn_label:
            eissn_span = eissn_label.find_next("span")
            eissn = eissn_span.get_text(strip=True) if eissn_span else ""

        extracted.append(
            {
                "publisher": "SAGE",
                "journal_name": journal_name,
                "journal_url": journal_url,
                "archive_page_url": archive_url,
                "issn": eissn,
                "pssn": pssn,
                "remarks": "Book" if is_book else "",
            }
        )

    return extracted


def start_process_show_publications(
    start_url: str = "https://journals.sagepub.com/action/showPublications?pageSize=10&startPage=0",
    base_url: str = "https://journals.sagepub.com",
    push_new_to_db: bool = False,
    include_books: bool = False,
    force_browser: bool = False,
):
    """
    Scrape SAGE journals via the official browse page:
      https://journals.sagepub.com/action/showPublications

    Pagination is done by following the 'next' anchor:
      <a title="next" class="next ... pagination__link" href="...startPage=N">
    """
    def _run_with_selenium():
        # Use undetected_chromedriver directly (avoids version_main pinning in common_code)
        import undetected_chromedriver as uc

        options = uc.ChromeOptions()
        # Keep it non-headless by default (less likely to be blocked)
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--start-maximized")

        driver = None
        try:
            driver = uc.Chrome(options=options)
            all_journals = []
            seen_page_urls = set()
            url_occurrence = {}
            page_no = 0
            total_expected = None

            next_url = start_url
            while next_url:
                page_no += 1
                print(f"Processing Page {page_no} -> {next_url}")
                driver.get(next_url)
                WebDriverWait(driver, 40).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                try:
                    WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "ul.search-result__publications"))
                    )
                except TimeoutException:
                    pass
                time.sleep(1)

                effective_url = (getattr(driver, "current_url", None) or next_url).strip()
                if effective_url in seen_page_urls:
                    print(f"(repeated page url -> stop) {effective_url}")
                    break
                seen_page_urls.add(effective_url)

                soup = BeautifulSoup(driver.page_source, "html.parser")

                if total_expected is None:
                    cnt = soup.find("span", class_=re.compile(r"\bresult__count\b"))
                    if cnt:
                        m = re.search(r"(\d+)", cnt.get_text(strip=True))
                        if m:
                            total_expected = int(m.group(1))

                extracted_this_page = _extract_show_publications_journals_from_soup(
                    soup,
                    base_url=base_url,
                    include_books=include_books,
                )
                if not extracted_this_page:
                    print("(0 journals extracted -> stop)")
                    break

                print(f"Collected {len(extracted_this_page)} journals on page {page_no}")
                for idx, j in enumerate(extracted_this_page, start=1):
                    jurl = (j.get("journal_url") or "").strip()
                    url_occurrence[jurl] = url_occurrence.get(jurl, 0) + 1
                    j["page_no"] = page_no
                    j["item_no"] = idx
                    j["occurrence"] = url_occurrence[jurl]
                    all_journals.append(j)

                next_a = soup.find("a", attrs={"title": "next"}, class_=re.compile(r"\bnext\b"))
                href = (next_a.get("href") or "").strip() if next_a else ""
                next_url = href if href else None

            expected_msg = f"expected {total_expected}" if total_expected is not None else "expected unknown"
            print(f"\nTotal Journals Extracted: {len(all_journals)} ({expected_msg})")
            return all_journals
        finally:
            if driver is not None:
                try:
                    driver.quit()
                except Exception:
                    pass

    if force_browser:
        all_journals = _run_with_selenium()
        # Save to Excel / push handling happens below; mimic request flow by continuing.
        total_expected = None
        expected_msg = f"expected {total_expected}" if total_expected is not None else "expected unknown"
        print(f"\nTotal Journals Extracted: {len(all_journals)} ({expected_msg})")
        _script_dir = os.path.dirname(os.path.abspath(__file__))
        output_excel = os.path.join(
            _script_dir, f"{DATABASE}_journals_check_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "SAGE Journals"
        headers = ["row_no", "page_no", "item_no", "occurrence", "publisher", "journal_name", "journal_url", "archive_page_url", "issn", "pssn", "remarks"]
        ws.append(headers)
        for row_no, j in enumerate(all_journals, start=1):
            ws.append([
                row_no,
                j.get("page_no", ""),
                j.get("item_no", ""),
                j.get("occurrence", ""),
                j.get("publisher", ""),
                j.get("journal_name", ""),
                j.get("journal_url", ""),
                j.get("archive_page_url", ""),
                j.get("issn", ""),
                j.get("pssn", ""),
                j.get("remarks", ""),
            ])
        wb.save(output_excel)
        print(f"Saved: {output_excel}")

        if push_new_to_db:
            # Push both journals and books (books will have remarks="Book" in Excel only)
            _compare_and_push_new(all_journals, DATABASE)
        return all_journals

    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/146.0.0.0 Safari/537.36"
            ),
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Connection": "keep-alive",
        }
    )

    all_journals = []
    seen_page_urls = set()
    url_occurrence = {}
    page_no = 0
    total_expected = None

    next_url = start_url
    while next_url:
        page_no += 1
        print(f"Processing Page {page_no} -> {next_url}")
        try:
            resp = session.get(next_url, timeout=90)
        except Exception as e:
            print(f"Request failed: {e}")
            break

        effective_url = (str(resp.url) if getattr(resp, "url", None) else next_url).strip()
        if effective_url in seen_page_urls:
            print(f"(repeated page url -> stop) {effective_url}")
            break
        seen_page_urls.add(effective_url)

        if not resp.ok:
            if resp.status_code == 403:
                print("HTTP 403 (blocked). Switching to browser-based extraction.")
                all_journals = _run_with_selenium()
                break
            print(f"HTTP {resp.status_code} on {effective_url}")
            break

        soup = BeautifulSoup(resp.text, "html.parser")

        if total_expected is None:
            cnt = soup.find("span", class_=re.compile(r"\bresult__count\b"))
            if cnt:
                m = re.search(r"(\d+)", cnt.get_text(strip=True))
                if m:
                    total_expected = int(m.group(1))

        extracted_this_page = _extract_show_publications_journals_from_soup(
            soup,
            base_url=base_url,
            include_books=include_books,
        )
        if not extracted_this_page:
            print("(0 journals extracted -> stop)")
            break

        print(f"Collected {len(extracted_this_page)} journals on page {page_no}")
        for idx, j in enumerate(extracted_this_page, start=1):
            jurl = (j.get("journal_url") or "").strip()
            url_occurrence[jurl] = url_occurrence.get(jurl, 0) + 1
            j["page_no"] = page_no
            j["item_no"] = idx
            j["occurrence"] = url_occurrence[jurl]
            all_journals.append(j)

        next_a = soup.find("a", attrs={"title": "next"}, class_=re.compile(r"\bnext\b"))
        href = (next_a.get("href") or "").strip() if next_a else ""
        next_url = href if href else None

    expected_msg = f"expected {total_expected}" if total_expected is not None else "expected unknown"
    print(f"\nTotal Journals Extracted: {len(all_journals)} ({expected_msg})")

    _script_dir = os.path.dirname(os.path.abspath(__file__))
    output_excel = os.path.join(
        _script_dir, f"{DATABASE}_journals_check_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "SAGE Journals"
    headers = ["row_no", "page_no", "item_no", "occurrence", "publisher", "journal_name", "journal_url", "archive_page_url", "issn", "pssn", "remarks"]
    ws.append(headers)
    for row_no, j in enumerate(all_journals, start=1):
        ws.append([
            row_no,
            j.get("page_no", ""),
            j.get("item_no", ""),
            j.get("occurrence", ""),
            j.get("publisher", ""),
            j.get("journal_name", ""),
            j.get("journal_url", ""),
            j.get("archive_page_url", ""),
            j.get("issn", ""),
            j.get("pssn", ""),
            j.get("remarks", ""),
        ])
    wb.save(output_excel)
    print(f"Saved: {output_excel}")

    if push_new_to_db:
        # Push both journals and books (books will have remarks="Book" in Excel only)
        _compare_and_push_new(all_journals, DATABASE)

    return all_journals


def process_heartbeat(process_id):
    while True:
        report_status(process_id, "running", "Process alive")
        time.sleep(300)  # 5 minutes

def _retry_skipped_links(SKIPPED_FILE, driver, links_only=False, links_queue_file=None):
    """
    Before the main journal loop, retry every URL in SKIPPED_FILE.

    --links-only mode  → for /toc/ issue pages: collect article links and append
                         them to links_queue_file (same format as the normal pass).
    full mode          → for /toc/ issue pages: collect links but skip writing to
                         queue (links will be visited normally in the main loop);
                         for /doi/ or /abs/ URLs: skip (handled by main loop).

    After each successful load the entry is removed from SKIPPED_FILE so it is
    never retried again.  Entries that still fail stay in the file.
    """
    if not os.path.exists(SKIPPED_FILE):
        return driver

    with open(SKIPPED_FILE, "r", encoding="utf-8") as f:
        entries = [json.loads(ln) for ln in f if ln.strip()]

    if not entries:
        return driver

    print(f"\n{'='*60}")
    print(f"[RETRY] {len(entries)} skipped entries found — retrying before main loop …")
    print(f"{'='*60}")

    still_skipped = []
    base_url = "https://journals.sagepub.com/"
    _break_at = len(entries)  # tracks where loop stopped if driver died

    for _i, entry in enumerate(entries):
        jid = entry.get("jid", "")
        url = (entry.get("url") or "").strip()
        reason = entry.get("reason", "")

        if not url:
            still_skipped.append(entry)
            continue

        print(f"\n  [RETRY] {url}  (reason: {reason})")

        try:
            # Use safe_get so CAPTCHA clicks and IP-block recovery are handled
            driver = safe_get(SKIPPED_FILE, driver, url, journal_id=jid)
            if driver is None:
                print(f"    ❌ Still blocked after safe_get — keeping in skipped file")
                still_skipped.append(entry)
                _break_at = _i + 1
                break  # driver is gone; stop retrying

            page_html = driver.page_source
            soup = BeautifulSoup(page_html, "html.parser")

            # ── Issue TOC page (/toc/...) ──────────────────────────────────
            if "/toc/" in url:
                # URL format: /toc/{sage_jid}/{vol}/{issue}
                path_parts = url.rstrip("/").split("/toc/", 1)[-1].split("/")
                vol       = path_parts[1] if len(path_parts) > 1 else ""
                issue_num = path_parts[2] if len(path_parts) > 2 else ""

                containers = soup.find_all("div", class_="issue-item__container")
                links_found = []
                seen = set()
                for container in containers:
                    for title_tag in container.find_all("div", class_="issue-item__title"):
                        a = title_tag.find("a", href=True)
                        if not a:
                            continue
                        article_url = urljoin(base_url, a["href"])
                        if not article_url or article_url in seen:
                            continue
                        seen.add(article_url)
                        title      = a.get_text(strip=True)
                        art_type   = ""
                        pub_date   = ""
                        header     = container.find("div", class_="issue-item__header")
                        if header:
                            spans = header.find_all("span")
                            if len(spans) >= 2:
                                art_type = spans[1].get_text(strip=True)
                            for sp in spans:
                                t = sp.get_text(strip=True)
                                if "First published" in t:
                                    pub_date = t.replace("First published", "").strip()
                                    break
                        pdf_url = ""
                        actions = container.find("div", class_="issue-item__actions")
                        if actions:
                            a_pdf = actions.find("a", title="download", href=True)
                            if a_pdf:
                                pdf_url = urljoin(base_url, a_pdf["href"])
                        links_found.append({
                            "jid": jid,
                            "article_url": article_url,
                            "article_title": title,
                            "article_type": art_type,
                            "published_date": pub_date,
                            "pdf": pdf_url,
                            "volume": str(vol),
                            "issue": str(issue_num),
                            "published_year": "",  # filled in below from published_date
                        })

                # Back-fill published_year from first article's published_date
                # e.g. "December 2024" or "1 January 2025" → "2024" / "2025"
                _yr_match = None
                for lnk in links_found:
                    m = re.search(r"\b(20\d{2}|19\d{2})\b", lnk.get("published_date", ""))
                    if m:
                        _yr_match = m.group(1)
                        break
                if _yr_match:
                    for lnk in links_found:
                        lnk["published_year"] = _yr_match

                if links_found:
                    if links_only and links_queue_file:
                        with open(links_queue_file, "a", encoding="utf-8") as lf:
                            for lnk in links_found:
                                lf.write(json.dumps(lnk, ensure_ascii=False) + "\n")
                    print(f"    ✅ Collected {len(links_found)} links — removed from skipped")
                else:
                    print(f"    ✅ Page loaded but 0 article links found — removed from skipped")
                # Either way: page loaded fine → remove entry

            # ── Journal listing or other pages ─────────────────────────────
            else:
                print(f"    ✅ Page loaded OK — removed from skipped")

        except Exception as exc:
            print(f"    ❌ Error: {exc} — keeping in skipped file")
            still_skipped.append(entry)

    # If driver died mid-loop, preserve any entries we never got to
    if _break_at < len(entries):
        still_skipped.extend(entries[_break_at:])

    # Rewrite skipped file with only the entries that still failed
    with open(SKIPPED_FILE, "w", encoding="utf-8") as f:
        for entry in still_skipped:
            f.write(json.dumps(entry, ensure_ascii=False) + "\n")

    removed = len(entries) - len(still_skipped)
    print(f"\n[RETRY] Finished: {removed} resolved, {len(still_skipped)} still skipped.\n")
    return driver


def main(part, links_only=False, retry_skipped=False):
    global process_id
    process_id = f"{DATABASE}_part_{part}"

    # 🔹 START heartbeat thread
    t = threading.Thread(
        target=process_heartbeat,
        args=(process_id,),
        daemon=True
    )
    t.start()

    # 🔹 Report process started
    report_status(process_id, "started", "Process started")
    driver = None
    try:
        driver = create_driver(head)
        # Anchor paths to this folder so parts work regardless of process cwd
        CACHE_FILE   = os.path.join(_HERE, f"{DATABASE}_part_{part}.txt")
        SKIPPED_FILE = os.path.join(_HERE, f"{DATABASE}_skipped{part}.txt")
        STATE_FILE   = os.path.join(_HERE, f"{DATABASE}_last{part}.json")

        global ARTICLE_FILE
        ARTICLE_FILE = os.path.join(_HERE, f"{DATABASE}_articles.txt")
        driver.get("https://journals.sagepub.com/doi/10.1177/20584601251387564")

        # ── Retry skipped links only (explicit --links-only skipped command) ──────
        if retry_skipped:
            _links_queue_for_retry = os.path.join(_HERE, f"{DATABASE}_links_queue_{part}.jsonl")
            driver = _retry_skipped_links(
                SKIPPED_FILE, driver,
                links_only=True,
                links_queue_file=_links_queue_for_retry,
            )
            print("[skipped] Done — links written to queue. Exiting (no last_state used).")
            return  # driver closed by finally block below
        # ─────────────────────────────────────────────────────────────────────

        if not os.path.exists(CACHE_FILE):
            ensure_sage_journals_split_into_6()
        if not os.path.exists(CACHE_FILE):
            print(f"❌ Cache file {CACHE_FILE} not found. Run ensure_sage_journals_split_into_6() or add sage_part_original.txt.")
            journals = []
        else:
            journals = load_journals_from_cache(CACHE_FILE)

        # Build/load extracted volume/issue pairs cache (speed-up). We'll never skip for 2025/2026.
        try:
            raw_links_path = ensure_sage_journals_in_article_links_cache()
            index_dir = ensure_sage_journals_in_article_links_index(raw_links_path)
        except Exception as e:
            print(f"[CACHE] journals_in_article_links cache/index unavailable: {e}. No extracted-issue skipping.")
            index_dir = ""

        last_state = load_last_state(STATE_FILE)

        if last_state:
            last_vol = last_state.get("last_volume")
            last_issue = last_state.get("last_issue")
            last_url = last_state.get("last_url")
            resume_journal_id = last_state.get("journal_id")

            start_volume = last_vol
            start_issue = last_issue
            start_url = last_url
        else:
            resume_journal_id = None
            start_volume = start_issue = start_url = None

        for journal in journals:
            jid = journal.get("jid")

            if resume_journal_id and jid != resume_journal_id:
                continue

            if index_dir:
                journal["_extracted_pairs"] = load_extracted_pairs_for_jid(index_dir, str(jid))
            else:
                journal["_extracted_pairs"] = set()

            _links_queue = os.path.join(_HERE, f"{DATABASE}_links_queue_{part}.jsonl") if links_only else None
            driver = process_articles(
                STATE_FILE,
                SKIPPED_FILE,
                [journal],
                start_volume,
                start_issue,
                start_url,
                driver,
                resume_journal=resume_journal_id,
                links_only=links_only,
                links_queue_file=_links_queue,
                links_state_file=STATE_FILE,
            )
            # sys.exit()
            # reset resume flags
            start_volume = start_issue = start_url = resume_journal_id = None

    except Exception as e:
        # 🔴 CRASH ALERT
        report_status(process_id, "crashed", str(e))
        raise
    finally:
        if driver is not None:
            try:
                force_close_driver(driver)
            except Exception:
                pass
# from system_monitor import start_system_monitor
if __name__ == "__main__":
    # Extract only journals to Excel (no DB, no article mining): python sage.py --journals-only
    if "--journals-only" in sys.argv or "-j" in sys.argv:
        # If an Excel path is provided, use that file instead of scraping:
        #   python sage.py --journals-only sage_journals_check_YYYYMMDD_HHMMSS.xlsx
        excel_arg = next((a for a in sys.argv[1:] if a.lower().endswith(".xlsx")), None)
        if excel_arg:
            journals_from_excel = _load_journals_from_excel(excel_arg)
            # Default behavior when Excel is provided: push from this Excel to DB (new-only)
            _compare_and_push_new(journals_from_excel, DATABASE)
            sys.exit(0)
        else:
            # journals.sagepub.com browse list – pagination via <a title="next" ...>
            SAGE_JOURNALS_URL = "https://journals.sagepub.com/action/showPublications?pageSize=10&startPage=0"
            start_process_show_publications(
                SAGE_JOURNALS_URL,
                base_url="https://journals.sagepub.com",
                push_new_to_db=True,
                include_books=True,
                force_browser=True,
            )
            sys.exit(0)
    # Push an Excel explicitly:
    #   python sage.py --push-excel path.xlsx
    if "--push-excel" in sys.argv:
        idx = sys.argv.index("--push-excel")
        excel_path = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else ""
        push_all = "--push-all" in sys.argv
        if not excel_path or not str(excel_path).lower().endswith(".xlsx"):
            raise SystemExit("Usage: python sage.py --push-excel <file.xlsx> [--push-all]")
        push_excel_to_db(excel_path, database=DATABASE, push_all=push_all)
        sys.exit(0)
    if "--skipped" in sys.argv or "-s" in sys.argv:
        workers = SKIPPED_FLOW_WORKERS
        if "--workers" in sys.argv:
            idx = sys.argv.index("--workers")
            if idx + 1 < len(sys.argv):
                try:
                    workers = int(sys.argv[idx + 1])
                except ValueError:
                    pass
        run_skipped_flow(workers=workers)
        sys.exit(0)
    if "--common-emails-json" in sys.argv:
        idx = sys.argv.index("--common-emails-json")
        json_path = sys.argv[idx + 1] if idx + 1 < len(sys.argv) else ""
        workers = 5
        print_details = "--print-details" in sys.argv
        if "--workers" in sys.argv:
            widx = sys.argv.index("--workers")
            if widx + 1 < len(sys.argv):
                try:
                    workers = int(sys.argv[widx + 1])
                except ValueError:
                    pass
        run_common_emails_json_flow(json_path, workers=workers, print_details=print_details)
        sys.exit(0)
    if len(sys.argv) > 1 and sys.argv[1] == "split":
        try:
            n_parts = int(sys.argv[2]) if len(sys.argv) > 2 else NUM_PARTS
        except ValueError:
            raise SystemExit("Usage: python sage.py split <N>  (e.g. python sage.py split 40)")
        split_journals_into_n_parts(n_parts)
        sys.exit(0)
    if "--data-only" in sys.argv:
        # Phase 2: enrich article links already collected in --links-only pass
        # Usage: python sage.py --data-only 1
        part = next((a for a in sys.argv[1:] if a.isdigit()), "1")
        queue_file      = os.path.join(_HERE, f"{DATABASE}_links_queue_{part}.jsonl")
        DATA_STATE_FILE = os.path.join(_HERE, f"{DATABASE}_last{part}_data.json")
        SKIPPED_FILE    = os.path.join(_HERE, f"{DATABASE}_skipped{part}.txt")
        driver = create_driver(head)
        try:
            enrich_article_links_with_data(DATA_STATE_FILE, SKIPPED_FILE, queue_file, driver=driver)
        finally:
            try:
                force_close_driver(driver)
            except Exception:
                pass
        sys.exit(0)
    else:
        # Run one part: python sage.py   or   python sage.py 1   or   python sage.py 3
        # With skipped retry: python sage.py 1 --links-only skipped
        _non_flag_args = [a for a in sys.argv[1:] if not a.startswith("-")]
        part = next((a for a in _non_flag_args if a.isdigit()), "1")
        links_only = "--links-only" in sys.argv
        # 'skipped' as a positional arg (not a digit) enables skipped-retry mode.
        # Only meaningful together with --links-only.
        retry_skipped = links_only and "skipped" in _non_flag_args
        main(part, links_only=links_only, retry_skipped=retry_skipped)
    # The below files are kept in main() method
    # CACHE_FILE = f"spring_part_{part}.txt"
    # SKIPPED_FILE = f"spring_skipped{part}.txt"
    # STATE_FILE = f"spring_last{part}.json"
# cd mining
# python sage_api.py
# for multiple files

# from multiprocessing import Process
# from springer_api import main

# if __name__ == "__main__":
#     processes = []

#     for n in [2, 3, 4, 5]:
#         p = Process(target=main, args=(n,))
#         p.start()
#         processes.append(p)
        
#     # wait for all to finish
#     for p in processes:
#         p.join()

#     print("✅ All processes finished.")
